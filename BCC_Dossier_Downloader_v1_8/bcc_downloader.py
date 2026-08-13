from __future__ import annotations

import base64
import csv
import mimetypes
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


ProgressFn = Callable[[str], None]


@dataclass(frozen=True)
class Customer:
    key: str
    name: str


@dataclass(frozen=True)
class Dossier:
    dossier_id: str
    label: str
    dossier_type: str


@dataclass(frozen=True)
class Architecture:
    section: str
    architecture_id: str
    label: str


@dataclass(frozen=True)
class Document:
    document_id: str
    architecture_id: str
    category: str
    title: str
    number: str
    date: str
    status: str


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\xa0", " ")).strip()


def normalize_match_text(value: str) -> str:
    value = clean_text(value).lower().replace("ё", "е")
    value = re.sub(r"[^0-9a-zа-я]+", " ", value, flags=re.I)
    return re.sub(r"\s+", " ", value).strip()


def decode_bcc_html(raw: bytes, content_type: str = "") -> str:
    """Decode BCC legacy HTML without losing Cyrillic characters.

    The ECD endpoints may omit/incorrectly declare charset. In that case browsers can
    turn Windows-1251 text into U+FFFD replacement characters. We fetch raw bytes and
    choose the best decoding locally.
    """
    if not raw:
        return ""
    m = re.search(r"charset\s*=\s*['\"]?([^;\s'\"]+)", content_type or "", re.I)
    candidates = []
    if m:
        candidates.append(m.group(1).strip())
    candidates += ["utf-8", "cp1251", "windows-1251"]
    best = None
    best_score = None
    seen = set()
    for enc in candidates:
        key = enc.lower()
        if key in seen:
            continue
        seen.add(key)
        try:
            text = raw.decode(enc, errors="replace")
        except LookupError:
            continue
        # Prefer decodings with fewer replacement chars and more Cyrillic letters.
        replacements = text.count("�")
        cyr = len(re.findall(r"[А-Яа-яЁё]", text))
        mojibake = text.count("Р") + text.count("С") if "Ð" in text or "Ñ" in text else 0
        score = (replacements * 1000) - cyr + mojibake
        if best_score is None or score < best_score:
            best_score = score
            best = text
    return best if best is not None else raw.decode("latin-1", errors="replace")


def normalize_dossier_number(value: str) -> str:
    value = clean_text(value).upper().replace(" ", "")
    return value.rstrip("/\\")


def dossier_matches(dossier: Dossier, requested_number: str) -> bool:
    if not requested_number.strip():
        return True
    wanted_num = normalize_dossier_number(requested_number)
    dossier_num = normalize_dossier_number(dossier.label)
    dossier_id = normalize_dossier_number(dossier.dossier_id)
    if wanted_num and wanted_num in {dossier_num, dossier_id}:
        return True
    wanted = normalize_match_text(requested_number)
    return wanted in {normalize_match_text(dossier.dossier_id), normalize_match_text(dossier.label)}


def _meaningful_document_number(value: str) -> bool:
    """Return True for a real-looking agreement/document number, not BCC placeholders."""
    raw = clean_text(value)
    norm = normalize_match_text(raw)
    if not norm or norm in {"-", "1", "0", "нет", "б н", "бн"}:
        return False
    # Real agreement numbers in BCC are typically several characters and contain digits,
    # often with '/', '-', letters, etc.  Keep this permissive because formats vary.
    return len(raw) >= 4 and bool(re.search(r"\d", raw))


def document_match_score(document: Document) -> tuple[str | None, int]:
    """Classify and rank a BCC document row.

    Business rules:
      * ``Примечание`` is the decisive field.
      * Acts/addenda/attachments are rejected even when the broad document category says
        "Договор ...".
      * Full contract wording has priority.
      * If full wording is absent, BCC abbreviations ``ДКП`` and ``ДФЛ`` are accepted as
        fallbacks when the surrounding category supports that interpretation.
      * A meaningful ``Номер документа`` increases confidence so rows such as
        ``643/BL/27-07`` beat placeholders such as ``1`` or ``-``.
    """
    note = normalize_match_text(document.title)
    category = normalize_match_text(document.category)

    # Derivative documents must never be treated as the principal contract.
    reject_terms = (
        "акт",
        "дополнительное соглашение",
        "доп соглашение",
        "допик",
        "допник",
        "приложение",
        "изменение к договору",
    )
    if any(term in note for term in reject_terms):
        return None, 0

    kind = None
    score = 0

    # Strong/full-name matches first.
    if "заявление о присоединении" in note and "договор лизинга" in note:
        kind, score = "leasing_contract", 120
    elif "договор финансового лизинга" in note:
        kind, score = "leasing_contract", 115
    elif "договор лизинга" in note:
        kind, score = "leasing_contract", 110
    elif "договор купли продажи" in note:
        kind, score = "purchase_contract", 120
    else:
        tokens = set(note.split())
        # Fallback abbreviations. Require category confirmation to avoid interpreting an
        # unrelated note containing a short acronym as a target contract.
        if "дфл" in tokens and "лизинг" in category:
            kind, score = "leasing_contract", 75
        elif "дкп" in tokens and ("купли продажи" in category or "дкп" in set(category.split())):
            kind, score = "purchase_contract", 75

    if kind is None:
        return None, 0

    if _meaningful_document_number(document.number):
        score += 25
        # Separator-rich/long IDs are very typical of the principal agreement row.
        if len(clean_text(document.number)) >= 7 or any(ch in document.number for ch in "/-"):
            score += 5
    elif clean_text(document.number) in {"", "-", "1", "0"}:
        score -= 5

    return kind, score


def document_kind(document: Document) -> str | None:
    return document_match_score(document)[0]


def select_best_contract_documents(documents: list[Document]) -> list[Document]:
    """Return at most one best leasing contract and one best purchase contract."""
    best: dict[str, tuple[int, Document]] = {}
    for document in documents:
        kind, score = document_match_score(document)
        if not kind:
            continue
        current = best.get(kind)
        if current is None or score > current[0]:
            best[kind] = (score, document)
        elif current is not None and score == current[0]:
            # Stable deterministic tie-break: prefer a meaningful document number, then id.
            cur_doc = current[1]
            new_key = (_meaningful_document_number(document.number), len(clean_text(document.number)), document.document_id)
            cur_key = (_meaningful_document_number(cur_doc.number), len(clean_text(cur_doc.number)), cur_doc.document_id)
            if new_key > cur_key:
                best[kind] = (score, document)

    ordered = []
    for kind in ("leasing_contract", "purchase_contract"):
        if kind in best:
            ordered.append(best[kind][1])
    return ordered

def is_target_document(document: Document) -> bool:
    return document_kind(document) is not None


def safe_name(value: str, max_len: int = 120) -> str:
    value = clean_text(value)
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value)
    value = value.rstrip(" .")
    return (value[:max_len].rstrip(" .") or "unnamed")


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    n = 2
    while True:
        candidate = path.with_name(f"{stem} ({n}){suffix}")
        if not candidate.exists():
            return candidate
        n += 1


def normalized_document_filename(document: Document, suffix: str) -> str:
    """Create a short Windows-safe filename while keeping number/date when available."""
    kind = document_kind(document)
    if kind == "leasing_contract":
        prefix = "01_Заявление о присоединении_Договор лизинга"
    elif kind == "purchase_contract":
        prefix = "02_Договор купли-продажи"
    else:
        prefix = document.title or f"document_{document.document_id}"

    parts = [prefix]
    if clean_text(document.number):
        parts.append(f"№_{clean_text(document.number)}")
    if clean_text(document.date):
        parts.append(clean_text(document.date))
    # Keep the whole filename comfortably below Windows path/name limits.
    stem = safe_name("_".join(parts), 155)
    suffix = suffix if suffix.startswith(".") else f".{suffix}"
    return f"{stem}{suffix}"


def parse_customers(html: str) -> list[Customer]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Customer] = []
    for radio in soup.select('input[name="r_pcust_shortname"]'):
        key = clean_text(radio.get("value", ""))
        tr = radio.find_parent("tr")
        tds = tr.find_all("td") if tr else []
        name = clean_text(tds[1].get_text(" ", strip=True)) if len(tds) > 1 else ""
        if key:
            out.append(Customer(key=key, name=name))
    return out


def select_customers(customers: list[Customer], requested_key: str = "") -> tuple[list[Customer], str | None]:
    """Resolve a BIN search result to the intended client row.

    If a key is supplied, matching is exact after trimming whitespace. If no key is
    supplied, a single customer is accepted automatically; multiple customers are
    treated as ambiguous so the caller can ask for Ключ клиента instead of guessing.
    """
    requested = clean_text(requested_key)
    if requested:
        matched = [c for c in customers if clean_text(c.key) == requested]
        if matched:
            return matched, None
        available = "; ".join(f"{c.name} ({c.key})" for c in customers)
        return [], f"Ключ клиента {requested} not found. Available: {available}"
    if len(customers) <= 1:
        return customers, None
    available = "; ".join(f"{c.name} ({c.key})" for c in customers)
    return [], f"multiple clients found; enter Ключ клиента. Available: {available}"


def parse_dossiers(html: str) -> list[Dossier]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Dossier] = []
    for radio in soup.select('input[name="r_pdossier"]'):
        did = clean_text(radio.get("value", ""))
        tr = radio.find_parent("tr")
        direct_tds = tr.find_all("td", recursive=False) if tr else []
        label = ""
        dossier_type = ""
        if len(direct_tds) >= 2:
            label = clean_text(direct_tds[1].get_text(" ", strip=True))
            label = re.sub(r"^Досье\s*№:\s*", "", label, flags=re.I)
        if len(direct_tds) >= 3:
            dossier_type = clean_text(direct_tds[2].get_text(" ", strip=True))
            dossier_type = re.sub(r"^Тип:\s*", "", dossier_type, flags=re.I)
        if not label:
            parent = radio.parent
            label = clean_text(parent.get_text(" ", strip=True)) if parent else ""
            label = re.sub(r"^Досье\s*№:\s*", "", label, flags=re.I)
        if did:
            out.append(Dossier(dossier_id=did, label=label, dossier_type=dossier_type))
    seen = set()
    result = []
    for d in out:
        if d.dossier_id not in seen:
            seen.add(d.dossier_id)
            result.append(d)
    return result


def parse_architectures(html: str) -> list[Architecture]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Architecture] = []
    for item in soup.select("ul.parent[section][parchitecture]"):
        section = clean_text(item.get("section", ""))
        aid = clean_text(item.get("parchitecture", ""))
        label = clean_text(item.get_text(" ", strip=True)).lstrip("*").strip()
        if section and aid:
            out.append(Architecture(section=section, architecture_id=aid, label=label))
    seen = set()
    result = []
    for a in out:
        key = (a.section, a.architecture_id)
        if key not in seen:
            seen.add(key)
            result.append(a)
    return result


def parse_documents(html: str, architecture: Architecture) -> list[Document]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Document] = []
    for img in soup.select('img[name="document_file"][id]'):
        doc_id = clean_text(img.get("id", ""))
        if not doc_id.isdigit():
            continue
        tr = img.find_parent("tr")
        tds = tr.find_all("td") if tr else []
        # Captured BCC table: category=3, note/title=4, number=7, date=9, last action=15, status=16.
        def td_text(i: int) -> str:
            return clean_text(tds[i].get_text(" ", strip=True)) if i < len(tds) else ""

        category = td_text(3) or architecture.label
        title = td_text(4) or category or f"document_{doc_id}"
        out.append(
            Document(
                document_id=doc_id,
                architecture_id=architecture.architecture_id,
                category=category,
                title=title,
                number=td_text(7),
                date=td_text(9),
                status=" / ".join(x for x in (td_text(15), td_text(16)) if x),
            )
        )
    return out


def filename_from_content_disposition(header: str) -> str | None:
    if not header:
        return None
    # RFC 5987 variant first.
    m = re.search(r"filename\*=UTF-8''([^;]+)", header, re.I)
    if m:
        from urllib.parse import unquote
        return Path(unquote(m.group(1).strip().strip('"'))).name
    m = re.search(r'filename\s*=\s*"?([^";]+)"?', header, re.I)
    if m:
        raw = m.group(1).strip()
        raw = raw.replace("\\", "/")
        return raw.split("/")[-1]
    return None


def extension_for_content_type(content_type: str) -> str:
    ctype = (content_type or "").split(";", 1)[0].strip().lower()
    common = {
        "application/pdf": ".pdf",
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/tiff": ".tif",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/msword": ".doc",
        "application/vnd.ms-excel": ".xls",
    }
    return common.get(ctype) or mimetypes.guess_extension(ctype) or ".bin"


def load_bins(path: Path) -> list[str]:
    ext = path.suffix.lower()
    values: list[str] = []
    if ext in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                s = re.sub(r"\D", "", str(value))
                if len(s) == 12:
                    values.append(s)
                    break
    elif ext == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.reader(f):
                for value in row:
                    s = re.sub(r"\D", "", value)
                    if len(s) == 12:
                        values.append(s)
                        break
    else:
        with path.open("r", encoding="utf-8-sig") as f:
            for line in f:
                s = re.sub(r"\D", "", line)
                if len(s) == 12:
                    values.append(s)
    # stable deduplication
    return list(dict.fromkeys(values))


class ExcelLog:
    HEADERS = [
        "Timestamp", "BIN", "Customer", "Customer Key", "Dossier ID", "Dossier", "Section",
        "Architecture ID", "Category", "Document ID", "Document title",
        "Document number", "Document date", "Status", "Result", "Saved path", "Error"
    ]
    SUMMARY_HEADERS = [
        "Timestamp", "BIN", "Customer", "Customer Key", "Dossier ID", "Dossier",
        "Leasing contract", "Purchase agreement", "Files downloaded", "Errors"
    ]

    def __init__(self, path: Path):
        self.path = path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Download log"
        self.ws.append(self.HEADERS)
        self.summary = self.wb.create_sheet("Case summary")
        self.summary.append(self.SUMMARY_HEADERS)
        self.wb.save(self.path)

    def add(self, **kwargs):
        row = [
            datetime.now().isoformat(timespec="seconds"),
            kwargs.get("bin", ""), kwargs.get("customer", ""), kwargs.get("customer_key", ""), kwargs.get("dossier_id", ""),
            kwargs.get("dossier", ""), kwargs.get("section", ""), kwargs.get("architecture_id", ""),
            kwargs.get("category", ""), kwargs.get("document_id", ""), kwargs.get("document_title", ""),
            kwargs.get("document_number", ""), kwargs.get("document_date", ""), kwargs.get("status", ""),
            kwargs.get("result", ""), kwargs.get("saved_path", ""), kwargs.get("error", "")
        ]
        self.ws.append(row)
        self.wb.save(self.path)

    def add_summary(self, **kwargs):
        self.summary.append([
            datetime.now().isoformat(timespec="seconds"),
            kwargs.get("bin", ""), kwargs.get("customer", ""), kwargs.get("customer_key", ""),
            kwargs.get("dossier_id", ""), kwargs.get("dossier", ""),
            kwargs.get("leasing_contract", "NOT FOUND"), kwargs.get("purchase_contract", "NOT FOUND"),
            kwargs.get("files_downloaded", 0), kwargs.get("errors", 0),
        ])
        self.wb.save(self.path)


class BCCDownloader:
    """Uses an authenticated Playwright BrowserContext for BCC ECD HTTP requests."""

    def __init__(self, context, page, output_root: Path, progress: ProgressFn = print, event: Callable[[dict], None] | None = None):
        self.context = context
        self.page = page
        self.output_root = output_root
        self.progress = progress
        self.event = event or (lambda payload: None)
        parsed = urlparse(page.url)
        self.origin = f"{parsed.scheme}://{parsed.netloc}"
        self.base = self.origin + "/ecd"
        self.access = "3"
        self.log = ExcelLog(output_root / "download_log.xlsx")

    def _browser_fetch(self, url: str, method: str = "GET", data: dict[str, str] | None = None, binary: bool = False):
        """Run fetch inside the authenticated BCC page so corporate/HTTP auth state is preserved."""
        payload = {"url": url, "method": method, "data": data or {}, "binary": binary}
        result = self.page.evaluate(
            """async ({url, method, data, binary}) => {
                const opts = {
                    method,
                    credentials: 'include',
                    cache: 'no-store',
                    headers: {}
                };
                if (method !== 'GET' && method !== 'HEAD') {
                    opts.headers['Content-Type'] = 'application/x-www-form-urlencoded; charset=UTF-8';
                    opts.body = new URLSearchParams(data).toString();
                }
                const r = await fetch(url, opts);
                const headers = {};
                for (const [k, v] of r.headers.entries()) headers[k.toLowerCase()] = v;
                if (binary) {
                    const bytes = new Uint8Array(await r.arrayBuffer());
                    let binaryString = '';
                    const chunk = 0x8000;
                    for (let i = 0; i < bytes.length; i += chunk) {
                        binaryString += String.fromCharCode(...bytes.subarray(i, i + chunk));
                    }
                    return {ok: r.ok, status: r.status, headers, body_b64: btoa(binaryString)};
                }
                return {ok: r.ok, status: r.status, headers, text: await r.text()};
            }""",
            payload,
        )
        return result

    def _post(self, endpoint: str, data: dict[str, str]) -> str:
        # Fetch raw bytes so BCC's legacy Cyrillic encoding is decoded correctly.
        result = self._browser_fetch(self.base + endpoint, method="POST", data=data, binary=True)
        if not result.get("ok"):
            raise RuntimeError(f"POST {endpoint} failed: HTTP {result.get('status')}")
        raw = base64.b64decode(result.get("body_b64", ""))
        headers = {k.lower(): v for k, v in (result.get("headers") or {}).items()}
        return decode_bcc_html(raw, headers.get("content-type", ""))

    def download_bins(self, bins: Iterable[str], dossier_number: str = "", customer_key: str = ""):
        bins = list(bins)
        self.event({"type": "batch_start", "total_cases": len(bins)})
        for index, bin_value in enumerate(bins, 1):
            self.event({"type": "case_start", "index": index, "total_cases": len(bins), "bin": bin_value, "dossier": dossier_number})
            self.progress(f"[{index}] BIN {bin_value}: searching...")
            try:
                result = self.download_bin(bin_value, dossier_number=dossier_number, customer_key=customer_key)
                self.event({"type": "case_complete", **result})
            except Exception as exc:
                self.progress(f"BIN {bin_value}: ERROR: {exc}")
                self.log.add(bin=bin_value, result="BIN ERROR", error=str(exc))
                self.event({"type": "case_complete", "bin": bin_value, "dossier": dossier_number, "leasing_contract": "ERROR", "purchase_contract": "ERROR", "files_downloaded": 0, "errors": 1})

    def download_bin(self, bin_value: str, dossier_number: str = "", customer_key: str = "") -> dict:
        bin_dir = self.output_root / safe_name(bin_value)
        total_downloaded = 0
        total_errors = 0
        leasing_status = "NOT FOUND"
        purchase_status = "NOT FOUND"
        summary_customer = ""
        summary_customer_key = customer_key
        summary_dossier_id = ""
        summary_dossier = dossier_number

        html = self._post("/pkg_w_e_dossier.p_main_table", {
            "p_cust_shortname": "",
            "p_dossier_number": "",
            "p_iin_bin": bin_value,
            "p_pcust_shortname": "",
            "p_sloan_agreement": "",
            "p_min_row_to_fetch": "0",
            "p_max_row_to_fetch": "50",
        })
        customers = parse_customers(html)
        if not customers:
            self.progress(f"BIN {bin_value}: no customer found")
            self.log.add(bin=bin_value, result="NOT FOUND")
            self.log.add_summary(bin=bin_value, dossier=dossier_number, leasing_contract=leasing_status, purchase_contract=purchase_status)
            return {"bin": bin_value, "dossier": dossier_number, "leasing_contract": leasing_status, "purchase_contract": purchase_status, "files_downloaded": 0, "errors": 0}

        customers, selection_error = select_customers(customers, customer_key)
        if selection_error:
            self.progress(f"BIN {bin_value}: {selection_error}")
            self.log.add(bin=bin_value, customer_key=customer_key, result="CLIENT SELECTION REQUIRED", error=selection_error)
            self.log.add_summary(bin=bin_value, customer_key=customer_key, dossier=dossier_number, leasing_contract=leasing_status, purchase_contract=purchase_status, errors=1)
            return {"bin": bin_value, "dossier": dossier_number, "leasing_contract": leasing_status, "purchase_contract": purchase_status, "files_downloaded": 0, "errors": 1}

        downloaded_ids: set[str] = set()
        for customer in customers:
            summary_customer = customer.name
            summary_customer_key = customer.key
            self.progress(f"BIN {bin_value}: customer {customer.name} ({customer.key})")
            dossier_html = self._post("/pkg_w_e_dossier.p_dossier_info", {
                "p_pcust_shortname": customer.key,
                "p_access": self.access,
            })
            dossiers = parse_dossiers(dossier_html)
            if dossier_number.strip():
                dossiers = [d for d in dossiers if dossier_matches(d, dossier_number)]
                if not dossiers:
                    self.progress(f"BIN {bin_value}: dossier № {dossier_number} not found")
                    self.log.add(bin=bin_value, customer=customer.name, customer_key=customer.key, dossier=dossier_number, result="DOSSIER NOT FOUND")
                    continue
                specific = [d for d in dossiers if "общие документы" not in normalize_match_text(d.dossier_type)]
                if specific:
                    dossiers = specific
                self.progress(f"BIN {bin_value}: dossier filter matched {len(dossiers)} dossier(s)")
            if not dossiers:
                self.log.add(bin=bin_value, customer=customer.name, customer_key=customer.key, result="NO DOSSIER")
                continue

            for dossier in dossiers:
                summary_dossier_id = dossier.dossier_id
                summary_dossier = dossier.label
                self.event({"type": "dossier", "bin": bin_value, "dossier": dossier.label})
                self.progress(f"  dossier {dossier.dossier_id}: {dossier.label}")
                self._post("/pkg_w_e_dossier.p_dossier_info", {
                    "p_pcust_shortname": customer.key,
                    "p_dossier": dossier.dossier_id,
                    "p_access": self.access,
                    "p_arh_status": "",
                })
                doc_info = self._post("/pkg_w_e_dossier.p_document_info", {
                    "p_dossier": dossier.dossier_id,
                    "p_pcust_shortname": customer.key,
                    "p_button": "change_dossier",
                    "p_access": self.access,
                    "p_arh_status": "",
                })
                architectures = parse_architectures(doc_info)
                if not architectures:
                    self.log.add(bin=bin_value, customer=customer.name, customer_key=customer.key, dossier_id=dossier.dossier_id, dossier=dossier.label, result="NO DOCUMENT CATEGORIES")
                    continue

                customer_folder = bin_dir / safe_name(f"Client_{customer.key}_{customer.name}", 90)
                dossier_folder = customer_folder / safe_name(f"Dossier_{dossier.dossier_id}_{dossier.label}", 90)
                dossier_folder.mkdir(parents=True, exist_ok=True)
                dossier_candidates: list[tuple[Architecture, Document]] = []
                for architecture in architectures:
                    try:
                        detail = self._post("/PKG_W_E_DOSSIER.P_DOCUMENT_INFO_DETAIL", {
                            "p_dossier": dossier.dossier_id,
                            "p_section": architecture.section,
                            "p_access": self.access,
                            "p_architecture": architecture.architecture_id,
                            "p_arh_status": "",
                            "p_user_cnt": "0",
                            "p_centerproject_write": "0",
                        })
                        documents = parse_documents(detail, architecture)
                    except Exception as exc:
                        total_errors += 1
                        self.event({"type": "error"})
                        self.log.add(bin=bin_value, customer=customer.name, customer_key=customer.key, dossier_id=dossier.dossier_id, dossier=dossier.label, section=architecture.section, architecture_id=architecture.architecture_id, category=architecture.label, result="CATEGORY ERROR", error=str(exc))
                        continue

                    if not documents:
                        continue
                    matched_here = [d for d in documents if is_target_document(d)]
                    if not matched_here:
                        titles = [clean_text(f"{d.category} | {d.title} | № {d.number}") for d in documents[:3]]
                        if titles:
                            self.progress(f"    no target match in {architecture.label}; sample: {' ; '.join(titles)}")
                        continue
                    dossier_candidates.extend((architecture, d) for d in matched_here)

                # Rank across the whole dossier and download only the strongest row per
                # business document type. This prevents acts/duplicate rows from producing
                # several files while still allowing ДКП/ДФЛ abbreviation fallbacks.
                chosen_docs = select_best_contract_documents([d for _, d in dossier_candidates])
                for document in chosen_docs:
                    architecture = next(a for a, d in dossier_candidates if d.document_id == document.document_id)
                    kind, score = document_match_score(document)
                    self.progress(
                        f"    selected {kind}: {document.title} | № {document.number or '-'} "
                        f"(match score {score})"
                    )
                    if document.document_id in downloaded_ids:
                        self.event({"type": "skipped"})
                        continue
                    downloaded_ids.add(document.document_id)
                    ok = self._download_document(bin_value, customer, dossier, architecture, document, dossier_folder)
                    if ok:
                        total_downloaded += 1
                        self.event({"type": "downloaded", "kind": kind})
                        if kind == "leasing_contract":
                            leasing_status = "FOUND"
                        elif kind == "purchase_contract":
                            purchase_status = "FOUND"
                    else:
                        total_errors += 1
                        self.event({"type": "error"})

        if leasing_status == "NOT FOUND":
            self.event({"type": "not_found", "kind": "leasing_contract"})
        if purchase_status == "NOT FOUND":
            self.event({"type": "not_found", "kind": "purchase_contract"})

        self.log.add_summary(
            bin=bin_value, customer=summary_customer, customer_key=summary_customer_key,
            dossier_id=summary_dossier_id, dossier=summary_dossier,
            leasing_contract=leasing_status, purchase_contract=purchase_status,
            files_downloaded=total_downloaded, errors=total_errors,
        )
        self.progress(
            f"BIN {bin_value}: completed; Leasing contract: {leasing_status}; "
            f"Purchase agreement: {purchase_status}; files downloaded: {total_downloaded}"
        )
        return {
            "bin": bin_value, "dossier": summary_dossier, "leasing_contract": leasing_status,
            "purchase_contract": purchase_status, "files_downloaded": total_downloaded, "errors": total_errors,
        }

    def _download_document(
        self, bin_value: str, customer: Customer, dossier: Dossier,
        architecture: Architecture, document: Document, folder: Path
    ) -> bool:
        try:
            from urllib.parse import urlencode
            url = self.base + "/pkg_w_e_dossier.p_look_document?" + urlencode({
                "p_code_file": document.document_id,
                "p_name": "document_file",
            })
            result = self._browser_fetch(url, method="GET", binary=True)
            if not result.get("ok"):
                raise RuntimeError(f"HTTP {result.get('status')}")
            headers = {k.lower(): v for k, v in (result.get("headers") or {}).items()}
            content_type = headers.get("content-type", "")
            body = base64.b64decode(result.get("body_b64", ""))
            server_name = filename_from_content_disposition(headers.get("content-disposition", ""))
            suffix = Path(server_name).suffix if server_name else ""
            if not suffix:
                suffix = extension_for_content_type(content_type)
            file_name = normalized_document_filename(document, suffix)
            target = unique_path(folder / file_name)

            # Guard against login/error HTML being accidentally saved as a PDF.
            if "application/pdf" in content_type.lower() and not body.startswith(b"%PDF-"):
                raise RuntimeError("Server said PDF but response does not start with %PDF-; session may have expired")
            if "text/html" in content_type.lower() and body.lstrip().startswith((b"<", b"<!")):
                raise RuntimeError("Received HTML instead of a document; session may have expired or access is denied")

            target.write_bytes(body)
            self.progress(f"    saved {target.name}")
            self.log.add(
                bin=bin_value, customer=customer.name, customer_key=customer.key, dossier_id=dossier.dossier_id,
                dossier=dossier.label, section=architecture.section,
                architecture_id=architecture.architecture_id, category=document.category,
                document_id=document.document_id, document_title=document.title,
                document_number=document.number, document_date=document.date, status=document.status,
                result="DOWNLOADED", saved_path=str(target)
            )
            return True
        except Exception as exc:
            self.progress(f"    document {document.document_id}: ERROR: {exc}")
            self.log.add(
                bin=bin_value, customer=customer.name, customer_key=customer.key, dossier_id=dossier.dossier_id,
                dossier=dossier.label, section=architecture.section,
                architecture_id=architecture.architecture_id, category=document.category,
                document_id=document.document_id, document_title=document.title,
                document_number=document.number, document_date=document.date, status=document.status,
                result="DOWNLOAD ERROR", error=str(exc)
            )
            return False
