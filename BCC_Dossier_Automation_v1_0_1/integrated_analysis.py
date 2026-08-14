from __future__ import annotations

import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook, Workbook

ProgressFn = Callable[[str], None]
EventFn = Callable[[dict], None]


@dataclass
class IntegratedAnalysisReport:
    bins_processed: int = 0
    pair_folders: int = 0
    unmatched_files: int = 0
    discovered: int = 0
    processed: int = 0
    failed: int = 0
    auto_approved: int = 0
    quarantined: int = 0
    manifests: list[Path] | None = None

    def __post_init__(self):
        if self.manifests is None:
            self.manifests = []


def configure_tesseract() -> None:
    """Configure OCR if Tesseract is installed; digital PDFs do not require it."""
    try:
        import pytesseract
    except Exception:
        return
    configured = os.environ.get("TESSERACT_CMD")
    candidates = [
        configured,
        shutil.which("tesseract"),
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            pytesseract.pytesseract.tesseract_cmd = str(candidate)
            return


def _clean_previous_gate(bin_root: Path) -> None:
    """Never leave stale production decisions from an older run."""
    for name in ("_PRODUCTION_READY", "_QUARANTINE"):
        path = bin_root / name
        if path.exists():
            shutil.rmtree(path, ignore_errors=True)
    manifest = bin_root / "production_manifest.xlsx"
    if manifest.exists():
        manifest.unlink()


def _manifest_counts(path: Path) -> tuple[int, int]:
    if not path.exists():
        return 0, 0
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Production Gate"] if "Production Gate" in wb.sheetnames else wb.active
        approved = quarantined = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            decision = str(row[0] or "")
            if decision == "AUTO_APPROVED":
                approved += 1
            elif decision == "QUARANTINED":
                quarantined += 1
        return approved, quarantined
    finally:
        wb.close()


def _count_pair_folders(bin_root: Path) -> int:
    return sum(1 for p in bin_root.rglob("*_LEASE_*") if p.is_dir() and re.match(r"^\d+_LEASE_", p.name, re.I))


def _count_unmatched(bin_root: Path) -> int:
    total = 0
    for folder in bin_root.rglob("_REVIEW_UNMATCHED"):
        if folder.is_dir():
            total += sum(1 for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in {".pdf", ".docx", ".png", ".jpg", ".jpeg", ".tif", ".tiff"})
    return total


def _write_aggregate_manifest(output_root: Path, records: list[dict]) -> Path:
    path = output_root / "integrated_production_manifest.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Integrated Production Gate"
    ws.append(["BIN/IIN", "AUTO_APPROVED", "QUARANTINED", "PAIR_FOLDERS", "UNMATCHED", "BIN_MANIFEST"])
    for record in records:
        ws.append([
            record.get("bin"), record.get("approved"), record.get("quarantined"),
            record.get("pairs"), record.get("unmatched"), record.get("manifest"),
        ])
    ws.freeze_panes = "A2"
    for col, width in {"A":18, "B":18, "C":18, "D":18, "E":18, "F":75}.items():
        ws.column_dimensions[col].width = width
    wb.save(path)
    return path


def analyze_download_tree(
    output_root: Path,
    bins: Iterable[str],
    *,
    analysis_mode: str = "auto",
    progress: ProgressFn = print,
    event: EventFn | None = None,
) -> IntegratedAnalysisReport:
    """Run v35 automatically on downloaded BCC trees, one BIN at a time.

    Only normal paired folders are discoverable by the embedded analyzer. Documents in
    ``_REVIEW_UNMATCHED`` are intentionally excluded from automatic production analysis.
    """
    from app.services.folder_batch import run_folder_batch

    event = event or (lambda payload: None)
    configure_tesseract()
    root = Path(output_root).expanduser().resolve()
    report = IntegratedAnalysisReport()
    aggregate_records: list[dict] = []

    for bin_value in list(dict.fromkeys(str(b).strip() for b in bins if str(b).strip())):
        bin_root = root / bin_value
        if not bin_root.exists():
            progress(f"Analyzer: BIN folder not found, skipping: {bin_value}")
            continue

        pairs = _count_pair_folders(bin_root)
        unmatched = _count_unmatched(bin_root)
        report.pair_folders += pairs
        report.unmatched_files += unmatched
        event({"type": "analysis_bin_start", "bin": bin_value, "pairs": pairs, "unmatched": unmatched})
        progress(f"Analyzer v35: BIN {bin_value} — pair folders: {pairs}; unmatched files: {unmatched}")

        _clean_previous_gate(bin_root)
        batch = run_folder_batch(
            bin_root,
            bin_value,
            recursive=True,
            analysis_mode=analysis_mode,
            save_individual=True,
            save_combined=False,
            progress=lambda text, b=bin_value: progress(f"  [{b}] {text}"),
        )
        report.bins_processed += 1
        report.discovered += int(batch.discovered)
        report.processed += len(batch.outputs)
        report.failed += int(batch.failed)

        manifest = bin_root / "production_manifest.xlsx"
        approved, quarantined = _manifest_counts(manifest)
        report.auto_approved += approved
        report.quarantined += quarantined
        if manifest.exists():
            report.manifests.append(manifest)
        aggregate_records.append({
            "bin": bin_value,
            "approved": approved,
            "quarantined": quarantined,
            "pairs": pairs,
            "unmatched": unmatched,
            "manifest": str(manifest) if manifest.exists() else "",
        })
        event({
            "type": "analysis_bin_complete", "bin": bin_value,
            "processed": len(batch.outputs), "failed": int(batch.failed),
            "approved": approved, "quarantined": quarantined,
            "pairs": pairs, "unmatched": unmatched,
        })

    if aggregate_records:
        aggregate = _write_aggregate_manifest(root, aggregate_records)
        progress(f"Integrated production manifest: {aggregate}")
    event({
        "type": "analysis_complete",
        "pairs": report.pair_folders,
        "unmatched": report.unmatched_files,
        "processed": report.processed,
        "failed": report.failed,
        "approved": report.auto_approved,
        "quarantined": report.quarantined,
    })
    return report
