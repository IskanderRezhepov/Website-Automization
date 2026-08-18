from __future__ import annotations

import hashlib
import re
from pathlib import Path
from shutil import copy2, move
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def decision_from_result(result: dict) -> tuple[str, str, dict]:
    review = ((result.get("analysis") or {}).get("work_review") or {})
    overall = str(review.get("overall") or "REVIEW_REQUIRED")
    counts = review.get("counts") or {}
    if overall == "READY" and int(counts.get("check") or 0) == 0 and int(counts.get("error") or 0) == 0:
        return "AUTO_APPROVED", "Автопроверка полностью пройдена", review
    return "QUARANTINED", str(review.get("label_ru") or "Требует проверки"), review


def _safe_component(value: str, limit: int = 28) -> str:
    digest = hashlib.sha1(str(value).encode("utf-8", errors="ignore")).hexdigest()[:8]
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "")).strip(" ._")
    cleaned = cleaned[:limit].rstrip(" ._") or "item"
    return f"{cleaned}_{digest}"


def _pair_component(root: Path, source: Path) -> str:
    """Return a short stable folder key without mirroring the long BCC tree."""
    try:
        rel = source.resolve().relative_to(root.resolve())
        parents = list(rel.parents)
        # rel.parents is nearest first; search path parts directly for the lease folder.
        for part in rel.parts[:-1]:
            if re.match(r"^\d+_LEASE_", part, re.I):
                return _safe_component(part, 30)
    except Exception:
        pass
    return _safe_component(source.parent.name, 30)


def _destination(root: Path, source: Path, bucket: str, output_name: str) -> Path:
    """Use a short flat routing tree that remains well below Windows MAX_PATH."""
    pair = _pair_component(root, source)
    name = _safe_component(Path(output_name).stem, 24) + Path(output_name).suffix
    destination = root / bucket / pair / name
    destination.parent.mkdir(parents=True, exist_ok=True)
    return destination


def route_output(root: Path, source: Path, output: Path, result: dict) -> dict:
    decision, reason, review = decision_from_result(result)
    bucket = "_PRODUCTION_READY" if decision == "AUTO_APPROVED" else "_QUARANTINE"
    destination = _destination(root, source, bucket, output.name)
    copy2(output, destination)
    return {
        "source": str(source),
        "workbook": str(output),
        "routed_workbook": str(destination),
        "decision": decision,
        "reason": reason,
        "pass": int((review.get("counts") or {}).get("pass") or 0),
        "check": int((review.get("counts") or {}).get("check") or 0),
        "error": int((review.get("counts") or {}).get("error") or 0),
    }


def route_failed_source(root: Path, source: Path, error: str) -> dict:
    """Quarantine a source that could not be analyzed (for example OCR unavailable)."""
    destination = _destination(root, source, "_QUARANTINE", "FAILED_" + source.name)
    try:
        copy2(source, destination)
        routed = str(destination)
    except Exception:
        routed = ""
    reason = f"ANALYSIS_FAILED: {error}"
    return {
        "source": str(source),
        "workbook": "",
        "routed_workbook": routed,
        "decision": "QUARANTINED",
        "reason": reason,
        "pass": 0,
        "check": 0,
        "error": 1,
    }


def _pair_key(root: Path, source_text: str) -> str:
    try:
        source = Path(source_text)
        rel = source.resolve().relative_to(root.resolve())
        for part in rel.parts[:-1]:
            if re.match(r"^\d+_LEASE_", part, re.I):
                return part
        return str(source.parent)
    except Exception:
        return str(Path(source_text).parent)


def enforce_pair_atomicity(root: Path, records: list[dict]) -> list[dict]:
    """A leasing pair is production-ready only when every document in it is approved.

    One OCR failure or validation quarantine blocks the whole pair. This prevents a
    standalone leasing application from reaching production while its connected DKP
    failed analysis.
    """
    groups: dict[str, list[dict]] = {}
    for rec in records:
        groups.setdefault(_pair_key(root, str(rec.get("source") or "")), []).append(rec)

    for _key, group in groups.items():
        if not any(rec.get("decision") != "AUTO_APPROVED" for rec in group):
            continue
        for rec in group:
            if rec.get("decision") != "AUTO_APPROVED":
                continue
            rec["decision"] = "QUARANTINED"
            rec["reason"] = "PAIR_BLOCKED: связанный документ не прошёл автоматическую проверку"
            rec["check"] = max(1, int(rec.get("check") or 0))
            old = Path(str(rec.get("routed_workbook") or ""))
            source = Path(str(rec.get("source") or ""))
            workbook = Path(str(rec.get("workbook") or ""))
            if workbook.exists():
                new = _destination(root, source, "_QUARANTINE", workbook.name)
                try:
                    if old.exists():
                        new.parent.mkdir(parents=True, exist_ok=True)
                        move(str(old), str(new))
                    else:
                        copy2(workbook, new)
                    rec["routed_workbook"] = str(new)
                except Exception:
                    pass
    return records


def write_manifest(root: Path, records: Iterable[dict]) -> Path:
    records = list(records)
    path = root / "production_manifest.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Gate"
    ws.append(["Decision", "Source", "Workbook", "Routed workbook", "PASS", "CHECK", "ERROR", "Reason"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAF7")
    for rec in records:
        ws.append([rec.get("decision"), rec.get("source"), rec.get("workbook"), rec.get("routed_workbook"), rec.get("pass"), rec.get("check"), rec.get("error"), rec.get("reason")])
        fill = "C6EFCE" if rec.get("decision") == "AUTO_APPROVED" else "FFC7CE"
        ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=fill)
    ws.freeze_panes = "A2"
    widths = {"A":18,"B":55,"C":55,"D":65,"E":10,"F":10,"G":10,"H":70}
    for col,w in widths.items():
        ws.column_dimensions[col].width=w
    wb.save(path)
    return path
