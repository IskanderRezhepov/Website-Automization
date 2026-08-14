from __future__ import annotations

import json
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule


def _json_default(value):
    """Convert internal numeric/date objects to stable JSON values."""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, set):
        return sorted(value)
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def save_json(data: dict, path: Path) -> None:
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2, default=_json_default),
        encoding='utf-8',
    )


def _excel_value(value):
    """Return a scalar value accepted by openpyxl."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, set):
        value = sorted(value, key=str)
    if isinstance(value, (list, tuple, set)):
        return "; ".join(str(_excel_value(item)) for item in value if item not in (None, ""))
    if isinstance(value, dict):
        return "; ".join(
            f"{key}: {_excel_value(item)}" for key, item in value.items()
            if item not in (None, "", [], {})
        )
    return str(value)


def _append_excel_row(sheet, values) -> None:
    """Append one row after converting every item to an Excel-safe scalar."""
    sheet.append([_excel_value(value) for value in values])


def _enable_filter(sheet, header_row: int = 1) -> None:
    """Enable Excel column filters on a non-empty rectangular table."""
    if sheet.max_row >= header_row and sheet.max_column > 0:
        sheet.auto_filter.ref = (
            f"A{header_row}:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
        )


def _export_asset_row_ok(row: dict) -> bool:
    """Last-line safety filter for equipment rows before they reach Excel."""
    if not isinstance(row, dict):
        return False
    blob = " ".join(str(v or "") for k, v in row.items() if any(t in str(k).casefold() for t in (
        "model", "модель", "equipment_type", "asset_type", "vehicle_type", "вид техник",
        "equipment_name", "наименование", "brand", "марка", "vin"
    )))
    low = re.sub(r"\s+", " ", blob).strip().casefold()
    if not low:
        return False
    bad = (
        "үлгісі комплектация", "модель комплектация", "техникасының суреті",
        "техниканың суреті", "технические характеристики", "техникалық сипаттама",
        "изображение техники", "фото техники",
    )
    if any(token in low for token in bad):
        return False
    return True


def _configure_print(sheet) -> None:
    """Make every worksheet printable without horizontal page fragmentation."""
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = (
        "landscape" if sheet.max_column > 8 else "portrait"
    )
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    sheet.print_options.horizontalCentered = False
    sheet.page_margins = PageMargins(
        left=.25, right=.25, top=.4, bottom=.4, header=.15, footer=.15
    )
    if sheet.max_row and sheet.max_column:
        sheet.print_area = f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
        sheet.print_title_rows = "1:1"


def _is_identifier_key(value: object) -> bool:
    key = str(value or "").lower()
    return (
        "iin" in key
        or "bin" in key
        or "iban" in key
        or "vin" in key
        or key.endswith("_number")
        or key in {"serial_number", "chassis_number", "engine_number"}
    )


def _safe_sheet_title(title: str, used: set[str]) -> str:
    title = re.sub(r'[\[\]\*\?/\\:]', ' - ', title)
    title = re.sub(r'\s+', ' ', title).strip()[:31] or 'Документ'
    original = title
    counter = 2
    while title in used:
        suffix = f' {counter}'
        title = original[:31-len(suffix)] + suffix
        counter += 1
    used.add(title)
    return title


def save_excel(data: dict, path: Path) -> None:
    # v32 export lock: specification recovery runs at the last possible moment,
    # after all reconciliation/sanitizer stages and immediately before rows are
    # written to Excel.  No later pipeline step can collapse a verified
    # continuation-page table back to one row.
    try:
        from .production_hardening import (
            recover_multiitem_specification_assets,
            recover_numbered_specification_assets,
        )
        recover_multiitem_specification_assets(data)
        recover_numbered_specification_assets(data)
        # v33 export lock: after generic multi-item recovery, recover a verified
        # one-row commercial specification as the final fallback. This protects
        # future layouts where an earlier technical-characteristics table was
        # parsed as equipment but the actual commercial row is still clear.
        from .release_sanitizer import _recover_single_commercial_spec, _sanitize_equipment
        for _document in data.get('documents', []):
            _recover_single_commercial_spec(_document)
            _sanitize_equipment(_document)
    except Exception:
        # Export must remain available even if an optional recovery rule fails;
        # the already validated result is still exported rather than crashing.
        pass
    analysis = data.setdefault("analysis", {})
    analysis["quality_pipeline_version"] = "35.0"
    analysis["export_lock_version"] = "35.0"
    from .work_review import build_work_review
    work_review = build_work_review(data)
    analysis["work_review"] = work_review

    wb = Workbook()
    summary = wb.active
    summary.title = 'Сводка'
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    warning_fill = PatternFill('solid', fgColor='FFF2CC')
    header_font = Font(bold=True)

    summary.append(['Показатель', 'Значение'])
    client = data.get("client", {})
    if client:
        summary.append(["Клиент", client.get("name") or "Не определён"])
        client_identifier = client.get("iin_bin")
        client_identifier_text = str(client_identifier or "").strip()
        summary.append([
            "ИИН/БИН клиента",
            client_identifier_text
            if client_identifier_text
            else str(client_identifier)
            if client_identifier not in (None, "")
            else "Не определён",
        ])
        summary.cell(summary.max_row, 2).number_format = "@"
        summary.cell(summary.max_row, 2).quotePrefix = True
        summary.append(["Роль клиента", "Не применимо" if str(client.get("name") or "").strip() == "Не применимо" else (client.get("role_label_ru") or client.get("role") or "Не определена")])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    compact_export = bool((data.get("analysis") or {}).get("compact_export"))
    folder_document_count = (data.get("analysis") or {}).get("folder_document_count")
    reconciliation_count = (data.get("analysis") or {}).get("documents_used_for_reconciliation")
    show_dossier_sheet = bool((data.get("analysis") or {}).get("show_dossier_sheet", not compact_export))
    summary.append(['Документов в текущем Excel', len(data['documents'])])
    quality_version = (data.get('analysis') or {}).get('quality_pipeline_version')
    if quality_version:
        summary.append(['Версия анализатора', quality_version])
    if folder_document_count:
        summary.append([(data.get('analysis') or {}).get('folder_document_count_label') or 'Документов этого клиента', folder_document_count])
    dossier = data.get('dossier') or {}
    document_warnings = []
    review_notes = []
    for document in data.get('documents', []):
        source_blocking = document.get('blocking_warnings')
        source_notes = document.get('review_notes')
        if source_blocking is None:
            source_blocking = [w for w in document.get('warnings', []) if not isinstance(w, dict) or str(w.get('severity') or 'high').casefold() in {'high', 'critical'}]
        if source_notes is None:
            source_notes = [w for w in document.get('warnings', []) if isinstance(w, dict) and str(w.get('severity') or '').casefold() not in {'high', 'critical'}]
        for target, source in ((document_warnings, source_blocking), (review_notes, source_notes)):
            for warning in source:
                if isinstance(warning, dict):
                    message = warning.get('message_ru') or warning.get('message') or warning.get('warning')
                    field_label = str(warning.get('field') or warning.get('label_ru') or '').strip()
                    if field_label and field_label.casefold() not in str(message or '').casefold():
                        message = f"{field_label}: {message}"
                else:
                    message = warning
                message = str(message or '').strip()
                if message and message not in target:
                    target.append(message)
    families = [str((document.get('analysis') or {}).get('asset_family') or '') for document in data.get('documents', [])]
    nonclient_registration = bool(families) and all(family == 'registration' for family in families)
    client_name_ok = bool(str(client.get('name') or '').strip())
    client_id_ok = bool(re.fullmatch(r'\d{12}', str(client.get('iin_bin') or '').strip()))
    if not nonclient_registration and not client_name_ok:
        message = 'Имя клиента не определено.'
        if message not in document_warnings:
            document_warnings.insert(0, message)
    if not nonclient_registration and not client_id_ok:
        message = 'ИИН/БИН клиента не определён или распознан некорректно.'
        if message not in document_warnings:
            document_warnings.insert(0, message)
    overall_label = work_review.get("label_ru") or "ТРЕБУЕТ РУЧНОЙ ПРОВЕРКИ"
    summary.append(['Готовность результата', overall_label])
    review_counts = work_review.get("counts") or {}
    summary.append(['Автопроверка: PASS', review_counts.get('pass', 0)])
    summary.append(['Автопроверка: CHECK', review_counts.get('check', 0)])
    summary.append(['Автопроверка: ERROR', review_counts.get('error', 0)])
    if document_warnings:
        summary.append(['Требует проверки', len(document_warnings)])
        for number, message in enumerate(document_warnings[:20], start=1):
            summary.append([f'Проверка {number}', message])
    if review_notes:
        summary.append(['Дополнительных примечаний', len(review_notes)])
        for number, message in enumerate(review_notes[:10], start=1):
            summary.append([f'Примечание {number}', message])
    if dossier and not compact_export:
        summary.append(['Междокументных совпадений', dossier.get('counts', {}).get('match', 0)])
        summary.append(['Междокументных расхождений', dossier.get('counts', {}).get('mismatch', 0)])
        summary.append(['Не удалось проверить', dossier.get('counts', {}).get('not_enough_data', 0)])
        review = dossier.get("review") or {}
        summary.append(["Полей требуют проверки", review.get("field_warnings", 0)])
        warning_breakdown = review.get("warning_breakdown") or {}
        summary.append(["Исправленных значений", warning_breakdown.get("corrected", 0)])
        summary.append(["Значений-кандидатов", warning_breakdown.get("candidate", 0)])
        summary.append(["Рассчитанных значений", warning_breakdown.get("calculated", 0)])
        summary.append(["Значений без прямой цитаты", warning_breakdown.get("without_quote", 0)])
        summary.append(["Документов требуют внимания", review.get("documents_requiring_attention", 0)])
        summary.append(["Неполных загруженных документов", review.get("field_incomplete_documents", 0)])
        summary.append(["Отсутствующих ожидаемых документов", review.get("missing_expected_documents", 0)])
        selected = dossier.get("selected_field_completeness") or {}
        if selected:
            summary.append(["Выбрано пользовательских полей", selected.get("requested_count", 0)])
            summary.append(["Выбранных полей найдено напрямую", selected.get("directly_extracted_count", 0)])
            summary.append(["Выбранных полей с предупреждением", selected.get("warning_count", 0)])
            summary.append(["Выбранных полей не найдено", selected.get("missing_count", 0)])
            summary.append(["Выбранных полей не распознано", selected.get("unrecognized_count", 0)])
            summary.append(["Выбранных полей не применимо", selected.get("not_applicable_count", 0)])
            summary.append(["Полнота по выбору, %", selected.get("percentage", 0)])
        financial = dossier.get('financial') or {}
        if financial.get('total') or financial.get('amount_fields_count'):
            summary.append(['Денежных показателей', financial.get('amount_fields_count', 0)])
            summary.append(['Арифметических проверок', financial.get('total', 0)])
            summary.append(['Арифметических расхождений', financial.get('mismatch', 0)])
            summary.append(['Максимальное расхождение, тенге', financial.get('largest_difference_kzt', 0)])

    insurance_tables = [
        table for document in data.get("documents", [])
        for table in document.get("tables", []) if table.get("name") == "insurance_rows"
    ]
    gps_tables = [
        table for document in data.get("documents", [])
        for table in document.get("tables", []) if table.get("name") == "gps_rows"
    ]
    if insurance_tables or gps_tables:
        summary.append([])
        summary.append(["Страхование и GPS", "Значение"])
        for cell in summary[summary.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        summary.append(["Страховых документов", sum(t.get("row_count", 0) for t in insurance_tables)])
        summary.append(["GPS-документов", sum(t.get("row_count", 0) for t in gps_tables)])
        expiring = sum(
            1 for table in insurance_tables for row in table.get("rows", [])
            if row.get("status") in {"Истёк", "Скоро заканчивается"}
        )
        review = dossier.get("review") or {}
        summary.append([
            "Требуют внимания",
            max(expiring, int(review.get("documents_requiring_attention") or 0)),
        ])

    equipment_tables = [
        (document, table)
        for document in data.get("documents", [])
        for table in document.get("tables", [])
        if table.get("name") == "asset_vin_rows"
    ]
    equipment = dossier.get("equipment") or {}
    if equipment_tables or equipment.get("document_record_count"):
        summary.append([])
        summary.append(["Техника", "Значение"])
        for cell in summary[summary.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        equipment_rows = [
            row for _document, table in equipment_tables
            for row in table.get("rows", [])
            if isinstance(row, dict) and _export_asset_row_ok(row)
        ]
        explicit_quantity = sum(
            int(row.get("quantity") or 0) for row in equipment_rows
            if str(row.get("quantity") or "").isdigit()
            and row.get("status") not in {"candidate", "rejected", "Требует проверки"}
            and 0 < int(row.get("quantity") or 0) <= 100
        )
        confirmed_vins = {
            str(row.get("vin")).strip().upper() for row in equipment_rows
            if row.get("vin")
            and row.get("status") not in {"candidate", "Требует проверки"}
            and "ТРЕБУЕТ ПРОВЕРКИ" not in str(row.get("vin")).upper()
        }
        review_vin_variants = set()
        vin_pattern = re.compile(r"\b[A-HJ-NPR-Z0-9]{17}\b", re.I)
        for row in equipment_rows:
            if row.get("status") in {"candidate", "Требует проверки"} or "ТРЕБУЕТ ПРОВЕРКИ" in str(row.get("vin") or "").upper():
                for value in row.values():
                    review_vin_variants.update(v.upper() for v in vin_pattern.findall(str(value or "")))
        summary.append(["Записей о технике в документах", len(equipment_rows) or equipment.get("document_record_count") or "Не определено"])
        summary.append(["Количество единиц техники", explicit_quantity or (len(confirmed_vins) if confirmed_vins else "Не определено")])
        summary.append(["Подтверждённых VIN", len(confirmed_vins)])
        if review_vin_variants:
            summary.append(["Вариантов VIN для проверки", len(review_vin_variants)])
        by_type = {}
        for row in equipment_rows:
            label = row.get("equipment_type") or "Не определено"
            by_type[label] = by_type.get(label, 0) + int(row.get("quantity") or 1)
        if not by_type:
            by_type = equipment.get("types") or {}
        for label, quantity in sorted(by_type.items()):
            summary.append([f"Вид техники: {label}", quantity])

    # v35 work-ready validation section: one compact block on Summary plus a
    # dedicated editable manual-review worksheet.  READY is impossible while
    # an ERROR exists; CHECK items remain explicitly visible for a reviewer.
    summary.append([])
    summary.append(["ПРОВЕРКА / VALIDATION", "Результат"])
    validation_header_row = summary.max_row
    for cell in summary[validation_header_row]:
        cell.fill = header_fill
        cell.font = header_font
    summary.append(["Итоговый статус", overall_label])
    summary.append(["PASS", review_counts.get('pass', 0)])
    summary.append(["CHECK", review_counts.get('check', 0)])
    summary.append(["ERROR", review_counts.get('error', 0)])
    review_last_row = max(2, len(work_review.get("checks", [])) + 1)
    manual_status_formula = (
        f'=IF(COUNTIF(\'Проверка\'!F2:F{review_last_row},"Исправить")>0,"НЕ ГОТОВО — требуется исправление",'
        f'IF(COUNTIFS(\'Проверка\'!A2:A{review_last_row},"ERROR",\'Проверка\'!F2:F{review_last_row},"<>Подтверждено",\'Проверка\'!F2:F{review_last_row},"<>Не применимо")>0,"НЕ ГОТОВО — есть непроверенные ERROR",'
        f'IF(COUNTIFS(\'Проверка\'!A2:A{review_last_row},"CHECK",\'Проверка\'!F2:F{review_last_row},"<>Подтверждено",\'Проверка\'!F2:F{review_last_row},"<>Не применимо")>0,"ТРЕБУЕТ РУЧНОЙ ПРОВЕРКИ","РУЧНАЯ ПРОВЕРКА ЗАВЕРШЕНА")))'
    )
    summary.append(["Статус ручной проверки", manual_status_formula])
    important_checks = [c for c in work_review.get("checks", []) if c.get("status") != "PASS"][:12]
    if not important_checks:
        summary.append(["Ручная проверка", "Критических расхождений и обязательных CHECK не обнаружено."])
    else:
        for idx, check in enumerate(important_checks, start=1):
            summary.append([f"{check.get('status')} {idx}: {check.get('title')}", check.get('detail')])
            fill = PatternFill('solid', fgColor='F4CCCC' if check.get('status') == 'ERROR' else 'FFF2CC')
            for cell in summary[summary.max_row]:
                cell.fill = fill

    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 42
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary.freeze_panes = "A2"

    used = {'Сводка'}

    review_sheet = wb.create_sheet('Проверка')
    review_sheet.append(['Статус', 'Проверка', 'Что обнаружено', 'Значение', 'Что сделать', 'Ручная проверка', 'Комментарий'])
    for cell in review_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    review_sheet.freeze_panes = 'A2'
    manual_dv = DataValidation(type='list', formula1='"Не проверено,Подтверждено,Исправить,Не применимо"', allow_blank=False)
    review_sheet.add_data_validation(manual_dv)
    for check in work_review.get('checks', []):
        review_sheet.append([
            check.get('status'), check.get('title'), check.get('detail'), check.get('value'),
            check.get('action'), 'Не проверено' if check.get('status') != 'PASS' else 'Не применимо', '',
        ])
        row_no = review_sheet.max_row
        manual_dv.add(review_sheet.cell(row_no, 6))
        status = check.get('status')
        fill = PatternFill('solid', fgColor='D9EAD3' if status == 'PASS' else 'FFF2CC' if status == 'CHECK' else 'F4CCCC')
        review_sheet.cell(row_no, 1).fill = fill
        review_sheet.cell(row_no, 1).font = Font(bold=True)
    review_sheet.append([])
    review_sheet.append(['Проверил', '', '', '', '', '', ''])
    review_sheet.append(['Дата проверки', '', '', '', '', '', ''])
    review_sheet.append([])
    review_sheet.append(['ИНСТРУКЦИЯ', 'ERROR = результат заблокирован; CHECK = нужна ручная сверка; PASS = автоматическая проверка пройдена. В столбце «Ручная проверка» выберите результат проверки и при необходимости добавьте комментарий. После подтверждения всех ERROR/CHECK строк поле «Статус ручной проверки» на листе «Сводка» изменится автоматически.'])
    review_sheet.merge_cells(start_row=review_sheet.max_row, start_column=2, end_row=review_sheet.max_row, end_column=7)
    for col, width in {'A':14,'B':42,'C':72,'D':38,'E':55,'F':20,'G':45}.items():
        review_sheet.column_dimensions[col].width = width
    for row in review_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    _enable_filter(review_sheet)
    used.add('Проверка')

    if dossier and show_dossier_sheet:
        dossier_sheet = wb.create_sheet('Сверка досье')
        dossier_sheet.append(['Категория', 'Проверка', 'Статус', 'Результат', 'Доказательства'])
        for cell in dossier_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        status_ru = {'match': 'Совпадает', 'mismatch': 'Расхождение', 'not_enough_data': 'Недостаточно данных'}
        missing_documents = [
            item for item in dossier.get("completeness", [])
            if item.get("status") == "referenced_not_uploaded"
        ]
        missing_labels = {str(item.get("label_ru") or "").casefold() for item in missing_documents}
        for check in dossier.get('checks', []):
            message_text = str(check.get("message") or "").casefold()
            if check.get("status") == "not_enough_data" and any(
                label and label in message_text for label in missing_labels
            ):
                continue
            evidence = '; '.join(
                f"{item.get('filename')} / {item.get('field')} = {item.get('value')}"
                for item in check.get('evidence', [])
            )
            dossier_sheet.append([
                check.get('category'), check.get('check'), status_ru.get(check.get('status'), check.get('status')),
                check.get('message'), evidence,
            ])
        if missing_documents:
            for item in missing_documents:
                dossier_sheet.append([
                    "Комплектность досье",
                    item.get("label_ru"),
                    "Не загружен",
                    ", ".join(item.get("referenced_numbers") or []),
                    item.get("message"),
                ])
        for column, width in {'A':22,'B':42,'C':20,'D':75,'E':100}.items():
            dossier_sheet.column_dimensions[column].width = width
        for row in dossier_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        dossier_sheet.freeze_panes = 'A2'
        _enable_filter(dossier_sheet)
        used.add('Сверка досье')
    status_labels_ru = {
        'extracted': 'Извлечено', 'confirmed': 'Подтверждено',
        'corrected': 'Исправлено автоматически', 'calculated': 'Рассчитано',
        'candidate': 'Требует проверки', 'rejected': 'Не найдено',
        'reviewed': 'Проверено', 'manual': 'Введено вручную',
    }
    for document in data['documents']:
        sheet = wb.create_sheet(_safe_sheet_title(document['document_type_label_ru'], used))
        if compact_export:
            sheet.append(['Поле', 'Значение', 'Страница', 'Надёжность', 'Статус', 'Цитата'])
        else:
            sheet.append(['Поле', 'Категория', 'Значение', 'Проверка', 'Сообщение проверки', 'Исходное значение', 'Нормализованное значение', 'Причина исправления', 'Источник восстановления', 'Страница', 'Метод', 'Уверенность', 'Статус', 'Проверено', 'Примечание', 'Цитата'])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sorted_fields = sorted(
            document['fields'],
            key=lambda item: (
                item.get('page') is None,
                item.get('page') if item.get('page') is not None else 10**9,
                item.get('label_ru', ''),
            ),
        )
        field_rows: dict[str, int] = {}
        for item in sorted_fields:
            raw_value = item['value']
            if item.get("status") == "rejected":
                value = "Не определено"
            elif item.get("status") == "candidate":
                # Explicit review markers such as conflicting VIN candidates
                # must remain visible. Hide only unvalidated OCR garbage.
                raw_text = str(raw_value or "")
                value = (
                    _excel_value(raw_value)
                    if raw_text.startswith("Требует проверки:")
                    else f"Кандидат: {_excel_value(raw_value)}"
                    if (
                        item.get("value_type") in {"derived", "calculated"}
                        or item.get("name") in {
                            "engine_displacement_cm3",
                            "repayment_method_conflict",
                            "subsidy_date_conflict",
                        }
                    )
                    else "Не определено"
                )
            else:
                value = raw_value
            if isinstance(value, (dict, list, tuple, set)):
                value = _excel_value(value)
            if _is_identifier_key(item.get("name")) and value not in (None, ""):
                value = str(value)
            original_value = item.get('original_value')
            if original_value is None and item.get("status") in {"candidate", "rejected"}:
                original_value = raw_value
            if isinstance(original_value, (dict, list, tuple, set)):
                original_value = _excel_value(original_value)
            validation = item.get("validation") or {}
            recovered_from = item.get("recovered_from")
            if isinstance(recovered_from, (dict, list, tuple, set)):
                recovered_from = _excel_value(recovered_from)
            locations = item.get("source_locations") or {}
            location_text = "; ".join(
                f"{candidate}: стр. {', '.join(str(loc.get('page')) for loc in found[:5])}"
                for candidate, found in locations.items()
                if found
            )
            quote_value = (str(item.get('quote') or '') + (' | Места: ' + location_text if location_text else '')) or None
            if compact_export:
                if item.get("status") in {"candidate", "rejected"} and value == "Не определено":
                    continue
                _append_excel_row(sheet, [
                    item['label_ru'], value, item.get('page'), item.get('confidence'),
                    (status_labels_ru.get(item.get('status'), item.get('status'))
                     if item.get('status') not in (None, '', 'extracted', 'confirmed') else ''), quote_value,
                ])
                if item.get('status') in {'candidate', 'rejected'}:
                    for cell in sheet[sheet.max_row]:
                        cell.fill = warning_fill
            else:
                _append_excel_row(sheet, [
                    item['label_ru'], item.get('category'), value,
                    "Корректно" if validation.get("valid") else "Требует проверки" if validation else None,
                    validation.get("message"),
                    item.get("raw_value", original_value), item.get("normalized_value", item.get("value")),
                    item.get("correction_reason"), recovered_from,
                    item.get('page'), item.get('extraction_method'),
                    item.get('confidence'), item.get('status'), item.get('reviewed_at'),
                    item.get('notes'), quote_value,
                ])
            field_rows.setdefault(str(item.get("name") or ""), sheet.max_row)
            value_column = 2 if compact_export else 3
            if str(item.get("name") or "").endswith("_kzt"):
                sheet.cell(sheet.max_row, value_column).number_format = '#,##0.00'
            elif _is_identifier_key(item.get("name")):
                # Preserve text semantics in the value and both audit
                # representations. Otherwise Excel/LibreOffice can display a
                # 12-digit IIN/BIN in scientific notation in columns F/G.
                for column_index in ((2,) if compact_export else (3, 6, 7)):
                    cell = sheet.cell(sheet.max_row, column_index)
                    if cell.value not in (None, ""):
                        cell.value = str(cell.value)
                    cell.number_format = "@"
                    cell.quotePrefix = True
        if compact_export:
            field_rows = {}
        financing_row = field_rows.get("financing_amount_kzt")
        asset_row = (
            field_rows.get("lease_asset_value_kzt")
            or field_rows.get("purchase_total_kzt")
            or field_rows.get("total_amount_kzt")
        )
        advance_row = field_rows.get("advance_payment_kzt")
        financing_item = next(
            (item for item in sorted_fields if item.get("name") == "financing_amount_kzt"),
            None,
        )
        if (
            financing_row and asset_row and advance_row and financing_item
            and financing_item.get("value_type") == "calculated"
            and financing_item.get("status") in {"confirmed", "calculated"}
        ):
            sheet.cell(financing_row, 3).value = f"=C{asset_row}-C{advance_row}"
            sheet.cell(financing_row, 3).number_format = '#,##0.00'
        commission_row = field_rows.get("arrangement_commission_kzt")
        commission_rate_row = field_rows.get("arrangement_commission_percent")
        commission_check_row = field_rows.get("arrangement_commission_check")
        commission_item = next(
            (
                item for item in sorted_fields
                if item.get("name") == "arrangement_commission_kzt"
            ),
            None,
        )
        if (
            commission_row and commission_rate_row and asset_row
            and commission_item
            and commission_item.get("status") in {"confirmed", "calculated"}
        ):
            sheet.cell(commission_row, 3).value = (
                f"=C{asset_row}*C{commission_rate_row}/100"
            )
            sheet.cell(commission_row, 3).number_format = "#,##0.00"
        if (
            commission_check_row and commission_row
            and commission_rate_row and asset_row
        ):
            sheet.cell(commission_check_row, 3).value = (
                f'=IF(ABS(C{commission_row}-'
                f'C{asset_row}*C{commission_rate_row}/100)<0.01,'
                f'"Сходится","Расхождение")'
            )
        gps_equipment_row = field_rows.get("gps_equipment_total_kzt")
        gps_annual_row = field_rows.get("gps_annual_fee_kzt")
        gps_monthly_row = field_rows.get("gps_monthly_fee_kzt")
        gps_period_row = field_rows.get("gps_subscription_period_months")
        gps_total_row = field_rows.get("gps_service_fee_kzt")
        if gps_annual_row and gps_monthly_row and gps_period_row:
            sheet.cell(gps_annual_row, 3).value = (
                f"=C{gps_monthly_row}*C{gps_period_row}"
            )
            sheet.cell(gps_annual_row, 3).number_format = "#,##0.00"
        if gps_total_row and gps_equipment_row and gps_annual_row:
            sheet.cell(gps_total_row, 3).value = (
                f"=C{gps_equipment_row}+C{gps_annual_row}"
            )
            sheet.cell(gps_total_row, 3).number_format = "#,##0.00"
        field_widths = ({'A':38,'B':48,'C':11,'D':13,'E':18,'F':90} if compact_export else {'A':35,'B':18,'C':42,'D':18,'E':48,'F':42,'G':42,'H':45,'I':55,'J':11,'K':14,'L':13,'M':16,'N':24,'O':38,'P':90})
        for column, width in field_widths.items():
            sheet.column_dimensions[column].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        sheet.freeze_panes = 'A2'
        _enable_filter(sheet)

        for table in document.get('tables', []):
            table_sheet = wb.create_sheet(_safe_sheet_title('Таблица - ' + table.get('label_ru', 'Данные'), used))
            columns = []
            seen_column_keys = set()
            seen_column_labels = set()
            for column in table.get('columns', []):
                key = str(column.get("key") or "")
                label = str(column.get("label_ru") or key).strip().casefold()
                if not key or key in seen_column_keys or label in seen_column_labels:
                    continue
                seen_column_keys.add(key)
                seen_column_labels.add(label)
                columns.append(column)
            table_sheet.append([column.get('label_ru', column.get('key')) for column in columns])
            for cell in table_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for row in table.get('rows', []):
                if table.get('name') == 'asset_vin_rows' and not _export_asset_row_ok(row):
                    continue
                _append_excel_row(
                    table_sheet,
                    [row.get(column.get('key')) for column in columns],
                )
            for column_index, column in enumerate(columns, start=1):
                if str(column.get("key") or "").endswith("_kzt"):
                    for row_index in range(2, table_sheet.max_row + 1):
                        table_sheet.cell(row_index, column_index).number_format = '#,##0.00'
                elif (
                    _is_identifier_key(column.get("key"))
                    or (
                        table.get("name") == "identifier_register_rows"
                        and column.get("key") == "value"
                    )
                ):
                    for row_index in range(2, table_sheet.max_row + 1):
                        cell = table_sheet.cell(row_index, column_index)
                        if cell.value not in (None, ""):
                            cell.value = str(cell.value)
                        cell.number_format = "@"
                        cell.quotePrefix = True
            if table.get("name") == "insurance_rows" and table_sheet.max_row >= 2:
                by_key = {
                    str(column.get("key")): index
                    for index, column in enumerate(columns, start=1)
                }
                start_col = by_key.get("start_date")
                end_col = by_key.get("end_date")
                status_col = by_key.get("status")
                days_col = by_key.get("days_remaining")
                for row_index in range(2, table_sheet.max_row + 1):
                    if end_col and days_col:
                        end_ref = table_sheet.cell(row_index, end_col).coordinate
                        table_sheet.cell(row_index, days_col).value = (
                            f'=IF({end_ref}="","",'
                            f'DATE(RIGHT({end_ref},4),MID({end_ref},4,2),'
                            f'LEFT({end_ref},2))-TODAY())'
                        )
                        table_sheet.cell(row_index, days_col).number_format = "0"
                    if end_col and status_col:
                        end_ref = table_sheet.cell(row_index, end_col).coordinate
                        start_ref = (
                            table_sheet.cell(row_index, start_col).coordinate
                            if start_col else None
                        )
                        before_start = (
                            f'IF(TODAY()<DATE(RIGHT({start_ref},4),'
                            f'MID({start_ref},4,2),LEFT({start_ref},2)),'
                            f'"Ожидает начала действия",'
                            if start_ref else ""
                        )
                        close_before_start = ")" if start_ref else ""
                        table_sheet.cell(row_index, status_col).value = (
                            f'=IF({end_ref}="","",'
                            f'{before_start}IF(DATE(RIGHT({end_ref},4),'
                            f'MID({end_ref},4,2),LEFT({end_ref},2))<TODAY(),'
                            f'"Истёк",IF(DATE(RIGHT({end_ref},4),'
                            f'MID({end_ref},4,2),LEFT({end_ref},2))-TODAY()<=30,'
                            f'"Скоро заканчивается","Действует")){close_before_start})'
                        )
            if table.get("name") == "gps_rows" and table_sheet.max_row >= 2:
                by_key = {
                    str(column.get("key")): index
                    for index, column in enumerate(columns, start=1)
                }
                monthly_col = by_key.get("monthly_fee_kzt")
                months_col = by_key.get("subscription_period_months")
                annual_col = by_key.get("annual_fee_kzt")
                equipment_col = by_key.get("equipment_total_kzt")
                service_col = by_key.get("service_fee_kzt")
                for row_index in range(2, table_sheet.max_row + 1):
                    if monthly_col and months_col and annual_col:
                        monthly_ref = table_sheet.cell(
                            row_index, monthly_col
                        ).coordinate
                        months_ref = table_sheet.cell(
                            row_index, months_col
                        ).coordinate
                        table_sheet.cell(row_index, annual_col).value = (
                            f'=IF(OR({monthly_ref}="",{months_ref}=""),"",'
                            f"{monthly_ref}*{months_ref})"
                        )
                        table_sheet.cell(
                            row_index, annual_col
                        ).number_format = "#,##0.00"
                    if equipment_col and annual_col and service_col:
                        equipment_ref = table_sheet.cell(
                            row_index, equipment_col
                        ).coordinate
                        annual_ref = table_sheet.cell(
                            row_index, annual_col
                        ).coordinate
                        table_sheet.cell(row_index, service_col).value = (
                            f'=IF(OR({equipment_ref}="",{annual_ref}=""),"",'
                            f"{equipment_ref}+{annual_ref})"
                        )
                        table_sheet.cell(
                            row_index, service_col
                        ).number_format = "#,##0.00"
            for column_index, column_cells in enumerate(
                table_sheet.columns, start=1
            ):
                letter = column_cells[0].column_letter
                max_len = max(len(str(cell.value or '')) for cell in column_cells)
                key = (
                    columns[column_index - 1].get("key")
                    if column_index <= len(columns)
                    else ""
                )
                minimum_width = 20 if _is_identifier_key(key) else 12
                table_sheet.column_dimensions[letter].width = min(
                    max(max_len + 2, minimum_width), 55
                )
            for row in table_sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            table_sheet.freeze_panes = 'A2'
            _enable_filter(table_sheet)

        if document.get('warnings') and not compact_export:
            warn_sheet = wb.create_sheet(_safe_sheet_title('Контроль - ' + document['document_type_label_ru'], used))
            warn_sheet.append(['Важность', 'Поле', 'Сообщение'])
            for cell in warn_sheet[1]:
                cell.fill = warning_fill
                cell.font = header_font
            for warning in document['warnings']:
                _append_excel_row(
                    warn_sheet,
                    [warning['severity'], warning['field'], warning['message']],
                )
            warn_sheet.column_dimensions['A'].width = 16
            warn_sheet.column_dimensions['B'].width = 35
            warn_sheet.column_dimensions['C'].width = 90
            warn_sheet.freeze_panes = 'A2'
            _enable_filter(warn_sheet)

        page_texts = document.get("page_texts") or []
        if page_texts and not compact_export:
            text_sheet = wb.create_sheet(
                _safe_sheet_title(
                    "Текст - " + document["document_type_label_ru"],
                    used,
                )
            )
            text_sheet.append(["Страница", "Метод чтения", "Прочитанный текст"])
            for cell in text_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            methods = {
                item.get("page"): item.get("method")
                for item in document.get("page_methods", [])
            }
            for page in page_texts:
                raw_text = str(page.get("text") or "")
                # Excel cells are limited to 32,767 characters. Preserve every
                # readable character by splitting an unusually long page into
                # deterministic chunks instead of silently truncating it.
                chunks = [
                    raw_text[index:index + 32000]
                    for index in range(0, len(raw_text), 32000)
                ] or [""]
                for chunk_index, chunk in enumerate(chunks, start=1):
                    page_label = str(page.get("page"))
                    if len(chunks) > 1:
                        page_label += f" (часть {chunk_index})"
                    _append_excel_row(
                        text_sheet,
                        [
                            page_label,
                            methods.get(page.get("page")),
                            chunk,
                        ],
                    )
                    text_sheet.row_dimensions[text_sheet.max_row].height = 90
            text_sheet.column_dimensions["A"].width = 16
            text_sheet.column_dimensions["B"].width = 18
            text_sheet.column_dimensions["C"].width = 100
            for row in text_sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            text_sheet.freeze_panes = "A2"
            _enable_filter(text_sheet)

    # Excel sometimes remembers the last generated worksheet as active.
    # Explicitly select the summary so every downloaded workbook opens there.
    wb.active = 0
    for index, worksheet in enumerate(wb.worksheets):
        worksheet.sheet_view.tabSelected = index == 0
    summary.sheet_view.selection[0].activeCell = "A1"
    summary.sheet_view.selection[0].sqref = "A1"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    for worksheet in wb.worksheets:
        _configure_print(worksheet)
    wb.save(path)
