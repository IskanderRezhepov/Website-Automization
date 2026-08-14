from __future__ import annotations

from pathlib import Path
from shutil import copy2
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def decision_from_result(result: dict) -> tuple[str, str, dict]:
    review = ((result.get("analysis") or {}).get("work_review") or {})
    overall = str(review.get("overall") or "REVIEW_REQUIRED")
    counts = review.get("counts") or {}
    if overall == "READY" and int(counts.get("check") or 0) == 0 and int(counts.get("error") or 0) == 0:
        return "AUTO_APPROVED", "Автопроверка полностью пройдена", review
    return "QUARANTINED", str(review.get("label_ru") or "Требует проверки"), review


def _mirror_destination(root: Path, source: Path, bucket: str, output_name: str) -> Path:
    try:
        rel_parent = source.parent.resolve().relative_to(root.resolve())
    except Exception:
        rel_parent = Path()
    destination = root / bucket / rel_parent / output_name
    destination.parent.mkdir(parents=True, exist_ok=True)
    return destination


def route_output(root: Path, source: Path, output: Path, result: dict) -> dict:
    decision, reason, review = decision_from_result(result)
    bucket = "_PRODUCTION_READY" if decision == "AUTO_APPROVED" else "_QUARANTINE"
    destination = _mirror_destination(root, source, bucket, output.name)
    copy2(output, destination)
    return {
        "source": str(source),
        "workbook": str(output),
        "routed_workbook": str(destination),
        "decision": decision,
        "reason": reason,
        "pass": int((review.get("counts") or {}).get("pass") or 0),
        "check": int((review.get("counts") or {}).get("check") or 0),
        "error": int((review.get("counts") or {}).get("error") or 0),
    }


def write_manifest(root: Path, records: Iterable[dict]) -> Path:
    records = list(records)
    path = root / "production_manifest.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Gate"
    ws.append(["Decision", "Source", "Workbook", "Routed workbook", "PASS", "CHECK", "ERROR", "Reason"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAF7")
    for rec in records:
        ws.append([rec.get("decision"), rec.get("source"), rec.get("workbook"), rec.get("routed_workbook"), rec.get("pass"), rec.get("check"), rec.get("error"), rec.get("reason")])
        fill = "C6EFCE" if rec.get("decision") == "AUTO_APPROVED" else "FFC7CE"
        ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=fill)
    ws.freeze_panes = "A2"
    widths = {"A":18,"B":55,"C":55,"D":65,"E":10,"F":10,"G":10,"H":55}
    for col,w in widths.items(): ws.column_dimensions[col].width=w
    wb.save(path)
    return path
