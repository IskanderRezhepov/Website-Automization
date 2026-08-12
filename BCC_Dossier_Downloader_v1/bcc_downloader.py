from __future__ import annotations

import csv
import mimetypes
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


ProgressFn = Callable[[str], None]


@dataclass(frozen=True)
class Customer:
    key: str
    name: str


@dataclass(frozen=True)
class Dossier:
    dossier_id: str
    label: str
    dossier_type: str


@dataclass(frozen=True)
class Architecture:
    section: str
    architecture_id: str
    label: str


@dataclass(frozen=True)
class Document:
    document_id: str
    architecture_id: str
    category: str
    title: str
    number: str
    date: str
    status: str


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\xa0", " ")).strip()


def safe_name(value: str, max_len: int = 120) -> str:
    value = clean_text(value)
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value)
    value = value.rstrip(" .")
    return (value[:max_len].rstrip(" .") or "unnamed")


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    n = 2
    while True:
        candidate = path.with_name(f"{stem} ({n}){suffix}")
        if not candidate.exists():
            return candidate
        n += 1


def parse_customers(html: str) -> list[Customer]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Customer] = []
    for radio in soup.select('input[name="r_pcust_shortname"]'):
        key = clean_text(radio.get("value", ""))
        tr = radio.find_parent("tr")
        tds = tr.find_all("td") if tr else []
        name = clean_text(tds[1].get_text(" ", strip=True)) if len(tds) > 1 else ""
        if key:
            out.append(Customer(key=key, name=name))
    return out


def parse_dossiers(html: str) -> list[Dossier]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Dossier] = []
    for radio in soup.select('input[name="r_pdossier"]'):
        did = clean_text(radio.get("value", ""))
        tr = radio.find_parent("tr")
        row_text = clean_text(tr.get_text(" ", strip=True)) if tr else ""
        m_label = re.search(r"Досье\s*№:\s*([^Т]+?)(?:\s+Тип:|$)", row_text, re.I)
        m_type = re.search(r"Тип:\s*(.+)$", row_text, re.I)
        label = clean_text(m_label.group(1)) if m_label else row_text
        dossier_type = clean_text(m_type.group(1)) if m_type else ""
        if did:
            out.append(Dossier(dossier_id=did, label=label, dossier_type=dossier_type))
    # preserve order, remove duplicates caused by malformed markup
    seen = set()
    result = []
    for d in out:
        if d.dossier_id not in seen:
            seen.add(d.dossier_id)
            result.append(d)
    return result


def parse_architectures(html: str) -> list[Architecture]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Architecture] = []
    for item in soup.select("ul.parent[section][parchitecture]"):
        section = clean_text(item.get("section", ""))
        aid = clean_text(item.get("parchitecture", ""))
        label = clean_text(item.get_text(" ", strip=True)).lstrip("*").strip()
        if section and aid:
            out.append(Architecture(section=section, architecture_id=aid, label=label))
    seen = set()
    result = []
    for a in out:
        key = (a.section, a.architecture_id)
        if key not in seen:
            seen.add(key)
            result.append(a)
    return result


def parse_documents(html: str, architecture: Architecture) -> list[Document]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Document] = []
    for img in soup.select('img[name="document_file"][id]'):
        doc_id = clean_text(img.get("id", ""))
        if not doc_id.isdigit():
            continue
        tr = img.find_parent("tr")
        tds = tr.find_all("td") if tr else []
        # Captured BCC table: category=3, note/title=4, number=7, date=9, last action=15, status=16.
        def td_text(i: int) -> str:
            return clean_text(tds[i].get_text(" ", strip=True)) if i < len(tds) else ""

        category = td_text(3) or architecture.label
        title = td_text(4) or category or f"document_{doc_id}"
        out.append(
            Document(
                document_id=doc_id,
                architecture_id=architecture.architecture_id,
                category=category,
                title=title,
                number=td_text(7),
                date=td_text(9),
                status=" / ".join(x for x in (td_text(15), td_text(16)) if x),
            )
        )
    return out


def filename_from_content_disposition(header: str) -> str | None:
    if not header:
        return None
    # RFC 5987 variant first.
    m = re.search(r"filename\*=UTF-8''([^;]+)", header, re.I)
    if m:
        from urllib.parse import unquote
        return Path(unquote(m.group(1).strip().strip('"'))).name
    m = re.search(r'filename\s*=\s*"?([^";]+)"?', header, re.I)
    if m:
        raw = m.group(1).strip()
        raw = raw.replace("\\", "/")
        return raw.split("/")[-1]
    return None


def extension_for_content_type(content_type: str) -> str:
    ctype = (content_type or "").split(";", 1)[0].strip().lower()
    common = {
        "application/pdf": ".pdf",
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/tiff": ".tif",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/msword": ".doc",
        "application/vnd.ms-excel": ".xls",
    }
    return common.get(ctype) or mimetypes.guess_extension(ctype) or ".bin"


def load_bins(path: Path) -> list[str]:
    ext = path.suffix.lower()
    values: list[str] = []
    if ext in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                s = re.sub(r"\D", "", str(value))
                if len(s) == 12:
                    values.append(s)
                    break
    elif ext == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.reader(f):
                for value in row:
                    s = re.sub(r"\D", "", value)
                    if len(s) == 12:
                        values.append(s)
                        break
    else:
        with path.open("r", encoding="utf-8-sig") as f:
            for line in f:
                s = re.sub(r"\D", "", line)
                if len(s) == 12:
                    values.append(s)
    # stable deduplication
    return list(dict.fromkeys(values))


class ExcelLog:
    HEADERS = [
        "Timestamp", "BIN", "Customer", "Dossier ID", "Dossier", "Section",
        "Architecture ID", "Category", "Document ID", "Document title",
        "Document number", "Document date", "Status", "Result", "Saved path", "Error"
    ]

    def __init__(self, path: Path):
        self.path = path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Download log"
        self.ws.append(self.HEADERS)
        self.wb.save(self.path)

    def add(self, **kwargs):
        row = [
            datetime.now().isoformat(timespec="seconds"),
            kwargs.get("bin", ""), kwargs.get("customer", ""), kwargs.get("dossier_id", ""),
            kwargs.get("dossier", ""), kwargs.get("section", ""), kwargs.get("architecture_id", ""),
            kwargs.get("category", ""), kwargs.get("document_id", ""), kwargs.get("document_title", ""),
            kwargs.get("document_number", ""), kwargs.get("document_date", ""), kwargs.get("status", ""),
            kwargs.get("result", ""), kwargs.get("saved_path", ""), kwargs.get("error", "")
        ]
        self.ws.append(row)
        self.wb.save(self.path)


class BCCDownloader:
    """Uses an authenticated Playwright BrowserContext for BCC ECD HTTP requests."""

    def __init__(self, context, page, output_root: Path, progress: ProgressFn = print):
        self.context = context
        self.page = page
        self.output_root = output_root
        self.progress = progress
        parsed = urlparse(page.url)
        self.origin = f"{parsed.scheme}://{parsed.netloc}"
        self.base = self.origin + "/ecd"
        self.access = "3"
        self.log = ExcelLog(output_root / "download_log.xlsx")

    def _post(self, endpoint: str, data: dict[str, str]) -> str:
        response = self.context.request.post(self.base + endpoint, form=data, timeout=90_000)
        if not response.ok:
            raise RuntimeError(f"POST {endpoint} failed: HTTP {response.status}")
        return response.text()

    def download_bins(self, bins: Iterable[str]):
        for index, bin_value in enumerate(bins, 1):
            self.progress(f"[{index}] BIN {bin_value}: searching...")
            try:
                self.download_bin(bin_value)
            except Exception as exc:
                self.progress(f"BIN {bin_value}: ERROR: {exc}")
                self.log.add(bin=bin_value, result="BIN ERROR", error=str(exc))

    def download_bin(self, bin_value: str):
        bin_dir = self.output_root / safe_name(bin_value)
        bin_dir.mkdir(parents=True, exist_ok=True)

        html = self._post("/pkg_w_e_dossier.p_main_table", {
            "p_cust_shortname": "",
            "p_dossier_number": "",
            "p_iin_bin": bin_value,
            "p_pcust_shortname": "",
            "p_sloan_agreement": "",
            "p_min_row_to_fetch": "0",
            "p_max_row_to_fetch": "50",
        })
        customers = parse_customers(html)
        if not customers:
            self.progress(f"BIN {bin_value}: no customer found")
            self.log.add(bin=bin_value, result="NOT FOUND")
            return

        downloaded_ids: set[str] = set()
        for customer in customers:
            self.progress(f"BIN {bin_value}: customer {customer.name} ({customer.key})")
            dossier_html = self._post("/pkg_w_e_dossier.p_dossier_info", {
                "p_pcust_shortname": customer.key,
                "p_access": self.access,
            })
            dossiers = parse_dossiers(dossier_html)
            if not dossiers:
                self.log.add(bin=bin_value, customer=customer.name, result="NO DOSSIER")
                continue

            for dossier in dossiers:
                self.progress(f"  dossier {dossier.dossier_id}: {dossier.label}")
                # Select dossier to mirror the site's normal state transition.
                self._post("/pkg_w_e_dossier.p_dossier_info", {
                    "p_pcust_shortname": customer.key,
                    "p_dossier": dossier.dossier_id,
                    "p_access": self.access,
                    "p_arh_status": "",
                })
                doc_info = self._post("/pkg_w_e_dossier.p_document_info", {
                    "p_dossier": dossier.dossier_id,
                    "p_pcust_shortname": customer.key,
                    "p_button": "change_dossier",
                    "p_access": self.access,
                    "p_arh_status": "",
                })
                architectures = parse_architectures(doc_info)
                if not architectures:
                    self.log.add(
                        bin=bin_value, customer=customer.name, dossier_id=dossier.dossier_id,
                        dossier=dossier.label, result="NO DOCUMENT CATEGORIES"
                    )
                    continue

                dossier_folder = bin_dir / safe_name(f"Dossier_{dossier.dossier_id}_{dossier.label}", 90)
                for architecture in architectures:
                    try:
                        detail = self._post("/PKG_W_E_DOSSIER.P_DOCUMENT_INFO_DETAIL", {
                            "p_dossier": dossier.dossier_id,
                            "p_section": architecture.section,
                            "p_access": self.access,
                            "p_architecture": architecture.architecture_id,
                            "p_arh_status": "",
                            "p_user_cnt": "0",
                            "p_centerproject_write": "0",
                        })
                        documents = parse_documents(detail, architecture)
                    except Exception as exc:
                        self.log.add(
                            bin=bin_value, customer=customer.name, dossier_id=dossier.dossier_id,
                            dossier=dossier.label, section=architecture.section,
                            architecture_id=architecture.architecture_id, category=architecture.label,
                            result="CATEGORY ERROR", error=str(exc)
                        )
                        continue

                    if not documents:
                        continue

                    category_folder = dossier_folder / safe_name(
                        f"S{architecture.section}_{architecture.architecture_id}_{architecture.label}", 100
                    )
                    category_folder.mkdir(parents=True, exist_ok=True)
                    for document in documents:
                        if document.document_id in downloaded_ids:
                            continue
                        downloaded_ids.add(document.document_id)
                        self._download_document(
                            bin_value, customer, dossier, architecture, document, category_folder
                        )

        self.progress(f"BIN {bin_value}: completed; {len(downloaded_ids)} unique document(s)")

    def _download_document(
        self, bin_value: str, customer: Customer, dossier: Dossier,
        architecture: Architecture, document: Document, folder: Path
    ):
        try:
            response = self.context.request.get(
                self.base + "/pkg_w_e_dossier.p_look_document",
                params={"p_code_file": document.document_id, "p_name": "document_file"},
                timeout=120_000,
            )
            if not response.ok:
                raise RuntimeError(f"HTTP {response.status}")
            headers = {k.lower(): v for k, v in response.headers.items()}
            content_type = headers.get("content-type", "")
            body = response.body()
            server_name = filename_from_content_disposition(headers.get("content-disposition", ""))
            suffix = Path(server_name).suffix if server_name else ""
            if not suffix:
                suffix = extension_for_content_type(content_type)
            base_name = safe_name(Path(server_name).stem if server_name else document.title, 135)
            file_name = f"{document.document_id}_{base_name}{suffix}"
            target = unique_path(folder / file_name)

            # Guard against login/error HTML being accidentally saved as a PDF.
            if "application/pdf" in content_type.lower() and not body.startswith(b"%PDF-"):
                raise RuntimeError("Server said PDF but response does not start with %PDF-; session may have expired")
            if "text/html" in content_type.lower() and body.lstrip().startswith((b"<", b"<!")):
                raise RuntimeError("Received HTML instead of a document; session may have expired or access is denied")

            target.write_bytes(body)
            self.progress(f"    saved {target.name}")
            self.log.add(
                bin=bin_value, customer=customer.name, dossier_id=dossier.dossier_id,
                dossier=dossier.label, section=architecture.section,
                architecture_id=architecture.architecture_id, category=document.category,
                document_id=document.document_id, document_title=document.title,
                document_number=document.number, document_date=document.date, status=document.status,
                result="DOWNLOADED", saved_path=str(target)
            )
        except Exception as exc:
            self.progress(f"    document {document.document_id}: ERROR: {exc}")
            self.log.add(
                bin=bin_value, customer=customer.name, dossier_id=dossier.dossier_id,
                dossier=dossier.label, section=architecture.section,
                architecture_id=architecture.architecture_id, category=document.category,
                document_id=document.document_id, document_title=document.title,
                document_number=document.number, document_date=document.date, status=document.status,
                result="DOWNLOAD ERROR", error=str(exc)
            )
