from __future__ import annotations

import base64
import csv
import mimetypes
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


ProgressFn = Callable[[str], None]


@dataclass(frozen=True)
class Customer:
    key: str
    name: str


@dataclass(frozen=True)
class Dossier:
    dossier_id: str
    label: str
    dossier_type: str


@dataclass(frozen=True)
class Architecture:
    section: str
    architecture_id: str
    label: str


@dataclass(frozen=True)
class Document:
    document_id: str
    architecture_id: str
    category: str
    title: str
    number: str
    date: str
    status: str


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\xa0", " ")).strip()


def normalize_match_text(value: str) -> str:
    value = clean_text(value).lower().replace("ё", "е")
    value = re.sub(r"[^0-9a-zа-я]+", " ", value, flags=re.I)
    return re.sub(r"\s+", " ", value).strip()


def decode_bcc_html(raw: bytes, content_type: str = "") -> str:
    """Decode BCC legacy HTML without losing Cyrillic characters.

    The ECD endpoints may omit/incorrectly declare charset. In that case browsers can
    turn Windows-1251 text into U+FFFD replacement characters. We fetch raw bytes and
    choose the best decoding locally.
    """
    if not raw:
        return ""
    m = re.search(r"charset\s*=\s*['\"]?([^;\s'\"]+)", content_type or "", re.I)
    candidates = []
    if m:
        candidates.append(m.group(1).strip())
    candidates += ["utf-8", "cp1251", "windows-1251"]
    best = None
    best_score = None
    seen = set()
    for enc in candidates:
        key = enc.lower()
        if key in seen:
            continue
        seen.add(key)
        try:
            text = raw.decode(enc, errors="replace")
        except LookupError:
            continue
        # Prefer decodings with fewer replacement chars and more Cyrillic letters.
        replacements = text.count("�")
        cyr = len(re.findall(r"[А-Яа-яЁё]", text))
        mojibake = text.count("Р") + text.count("С") if "Ð" in text or "Ñ" in text else 0
        score = (replacements * 1000) - cyr + mojibake
        if best_score is None or score < best_score:
            best_score = score
            best = text
    return best if best is not None else raw.decode("latin-1", errors="replace")


def normalize_dossier_number(value: str) -> str:
    value = clean_text(value).upper().replace(" ", "")
    return value.rstrip("/\\")


def dossier_matches(dossier: Dossier, requested_number: str) -> bool:
    if not requested_number.strip():
        return True
    wanted_num = normalize_dossier_number(requested_number)
    dossier_num = normalize_dossier_number(dossier.label)
    if wanted_num and dossier_num:
        return wanted_num == dossier_num
    wanted = normalize_match_text(requested_number)
    return wanted in {normalize_match_text(dossier.dossier_id), normalize_match_text(dossier.label)}


def is_target_document(document: Document) -> bool:
    text = normalize_match_text(" ".join([document.category, document.title]))
    leasing_application = "заявление о присоединении" in text and "договор лизинга" in text
    sale_contract = "договор купли продажи" in text
    return leasing_application or sale_contract


def safe_name(value: str, max_len: int = 120) -> str:
    value = clean_text(value)
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value)
    value = value.rstrip(" .")
    return (value[:max_len].rstrip(" .") or "unnamed")


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    n = 2
    while True:
        candidate = path.with_name(f"{stem} ({n}){suffix}")
        if not candidate.exists():
            return candidate
        n += 1


def parse_customers(html: str) -> list[Customer]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Customer] = []
    for radio in soup.select('input[name="r_pcust_shortname"]'):
        key = clean_text(radio.get("value", ""))
        tr = radio.find_parent("tr")
        tds = tr.find_all("td") if tr else []
        name = clean_text(tds[1].get_text(" ", strip=True)) if len(tds) > 1 else ""
        if key:
            out.append(Customer(key=key, name=name))
    return out


def parse_dossiers(html: str) -> list[Dossier]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Dossier] = []
    for radio in soup.select('input[name="r_pdossier"]'):
        did = clean_text(radio.get("value", ""))
        tr = radio.find_parent("tr")
        direct_tds = tr.find_all("td", recursive=False) if tr else []
        label = ""
        dossier_type = ""
        if len(direct_tds) >= 2:
            label = clean_text(direct_tds[1].get_text(" ", strip=True))
            label = re.sub(r"^Досье\s*№:\s*", "", label, flags=re.I)
        if len(direct_tds) >= 3:
            dossier_type = clean_text(direct_tds[2].get_text(" ", strip=True))
            dossier_type = re.sub(r"^Тип:\s*", "", dossier_type, flags=re.I)
        if not label:
            parent = radio.parent
            label = clean_text(parent.get_text(" ", strip=True)) if parent else ""
            label = re.sub(r"^Досье\s*№:\s*", "", label, flags=re.I)
        if did:
            out.append(Dossier(dossier_id=did, label=label, dossier_type=dossier_type))
    seen = set()
    result = []
    for d in out:
        if d.dossier_id not in seen:
            seen.add(d.dossier_id)
            result.append(d)
    return result


def parse_architectures(html: str) -> list[Architecture]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Architecture] = []
    for item in soup.select("ul.parent[section][parchitecture]"):
        section = clean_text(item.get("section", ""))
        aid = clean_text(item.get("parchitecture", ""))
        label = clean_text(item.get_text(" ", strip=True)).lstrip("*").strip()
        if section and aid:
            out.append(Architecture(section=section, architecture_id=aid, label=label))
    seen = set()
    result = []
    for a in out:
        key = (a.section, a.architecture_id)
        if key not in seen:
            seen.add(key)
            result.append(a)
    return result


def parse_documents(html: str, architecture: Architecture) -> list[Document]:
    soup = BeautifulSoup(html, "html.parser")
    out: list[Document] = []
    for img in soup.select('img[name="document_file"][id]'):
        doc_id = clean_text(img.get("id", ""))
        if not doc_id.isdigit():
            continue
        tr = img.find_parent("tr")
        tds = tr.find_all("td") if tr else []
        # Captured BCC table: category=3, note/title=4, number=7, date=9, last action=15, status=16.
        def td_text(i: int) -> str:
            return clean_text(tds[i].get_text(" ", strip=True)) if i < len(tds) else ""

        category = td_text(3) or architecture.label
        title = td_text(4) or category or f"document_{doc_id}"
        out.append(
            Document(
                document_id=doc_id,
                architecture_id=architecture.architecture_id,
                category=category,
                title=title,
                number=td_text(7),
                date=td_text(9),
                status=" / ".join(x for x in (td_text(15), td_text(16)) if x),
            )
        )
    return out


def filename_from_content_disposition(header: str) -> str | None:
    if not header:
        return None
    # RFC 5987 variant first.
    m = re.search(r"filename\*=UTF-8''([^;]+)", header, re.I)
    if m:
        from urllib.parse import unquote
        return Path(unquote(m.group(1).strip().strip('"'))).name
    m = re.search(r'filename\s*=\s*"?([^";]+)"?', header, re.I)
    if m:
        raw = m.group(1).strip()
        raw = raw.replace("\\", "/")
        return raw.split("/")[-1]
    return None


def extension_for_content_type(content_type: str) -> str:
    ctype = (content_type or "").split(";", 1)[0].strip().lower()
    common = {
        "application/pdf": ".pdf",
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/tiff": ".tif",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/msword": ".doc",
        "application/vnd.ms-excel": ".xls",
    }
    return common.get(ctype) or mimetypes.guess_extension(ctype) or ".bin"


def load_bins(path: Path) -> list[str]:
    ext = path.suffix.lower()
    values: list[str] = []
    if ext in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                s = re.sub(r"\D", "", str(value))
                if len(s) == 12:
                    values.append(s)
                    break
    elif ext == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.reader(f):
                for value in row:
                    s = re.sub(r"\D", "", value)
                    if len(s) == 12:
                        values.append(s)
                        break
    else:
        with path.open("r", encoding="utf-8-sig") as f:
            for line in f:
                s = re.sub(r"\D", "", line)
                if len(s) == 12:
                    values.append(s)
    # stable deduplication
    return list(dict.fromkeys(values))


class ExcelLog:
    HEADERS = [
        "Timestamp", "BIN", "Customer", "Dossier ID", "Dossier", "Section",
        "Architecture ID", "Category", "Document ID", "Document title",
        "Document number", "Document date", "Status", "Result", "Saved path", "Error"
    ]

    def __init__(self, path: Path):
        self.path = path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Download log"
        self.ws.append(self.HEADERS)
        self.wb.save(self.path)

    def add(self, **kwargs):
        row = [
            datetime.now().isoformat(timespec="seconds"),
            kwargs.get("bin", ""), kwargs.get("customer", ""), kwargs.get("dossier_id", ""),
            kwargs.get("dossier", ""), kwargs.get("section", ""), kwargs.get("architecture_id", ""),
            kwargs.get("category", ""), kwargs.get("document_id", ""), kwargs.get("document_title", ""),
            kwargs.get("document_number", ""), kwargs.get("document_date", ""), kwargs.get("status", ""),
            kwargs.get("result", ""), kwargs.get("saved_path", ""), kwargs.get("error", "")
        ]
        self.ws.append(row)
        self.wb.save(self.path)


class BCCDownloader:
    """Uses an authenticated Playwright BrowserContext for BCC ECD HTTP requests."""

    def __init__(self, context, page, output_root: Path, progress: ProgressFn = print):
        self.context = context
        self.page = page
        self.output_root = output_root
        self.progress = progress
        parsed = urlparse(page.url)
        self.origin = f"{parsed.scheme}://{parsed.netloc}"
        self.base = self.origin + "/ecd"
        self.access = "3"
        self.log = ExcelLog(output_root / "download_log.xlsx")

    def _browser_fetch(self, url: str, method: str = "GET", data: dict[str, str] | None = None, binary: bool = False):
        """Run fetch inside the authenticated BCC page so corporate/HTTP auth state is preserved."""
        payload = {"url": url, "method": method, "data": data or {}, "binary": binary}
        result = self.page.evaluate(
            """async ({url, method, data, binary}) => {
                const opts = {
                    method,
                    credentials: 'include',
                    cache: 'no-store',
                    headers: {}
                };
                if (method !== 'GET' && method !== 'HEAD') {
                    opts.headers['Content-Type'] = 'application/x-www-form-urlencoded; charset=UTF-8';
                    opts.body = new URLSearchParams(data).toString();
                }
                const r = await fetch(url, opts);
                const headers = {};
                for (const [k, v] of r.headers.entries()) headers[k.toLowerCase()] = v;
                if (binary) {
                    const bytes = new Uint8Array(await r.arrayBuffer());
                    let binaryString = '';
                    const chunk = 0x8000;
                    for (let i = 0; i < bytes.length; i += chunk) {
                        binaryString += String.fromCharCode(...bytes.subarray(i, i + chunk));
                    }
                    return {ok: r.ok, status: r.status, headers, body_b64: btoa(binaryString)};
                }
                return {ok: r.ok, status: r.status, headers, text: await r.text()};
            }""",
            payload,
        )
        return result

    def _post(self, endpoint: str, data: dict[str, str]) -> str:
        # Fetch raw bytes so BCC's legacy Cyrillic encoding is decoded correctly.
        result = self._browser_fetch(self.base + endpoint, method="POST", data=data, binary=True)
        if not result.get("ok"):
            raise RuntimeError(f"POST {endpoint} failed: HTTP {result.get('status')}")
        raw = base64.b64decode(result.get("body_b64", ""))
        headers = {k.lower(): v for k, v in (result.get("headers") or {}).items()}
        return decode_bcc_html(raw, headers.get("content-type", ""))

    def download_bins(self, bins: Iterable[str], dossier_number: str = ""):
        for index, bin_value in enumerate(bins, 1):
            self.progress(f"[{index}] BIN {bin_value}: searching...")
            try:
                self.download_bin(bin_value, dossier_number=dossier_number)
            except Exception as exc:
                self.progress(f"BIN {bin_value}: ERROR: {exc}")
                self.log.add(bin=bin_value, result="BIN ERROR", error=str(exc))

    def download_bin(self, bin_value: str, dossier_number: str = ""):
        bin_dir = self.output_root / safe_name(bin_value)
        bin_dir.mkdir(parents=True, exist_ok=True)

        html = self._post("/pkg_w_e_dossier.p_main_table", {
            "p_cust_shortname": "",
            "p_dossier_number": "",
            "p_iin_bin": bin_value,
            "p_pcust_shortname": "",
            "p_sloan_agreement": "",
            "p_min_row_to_fetch": "0",
            "p_max_row_to_fetch": "50",
        })
        customers = parse_customers(html)
        if not customers:
            self.progress(f"BIN {bin_value}: no customer found")
            self.log.add(bin=bin_value, result="NOT FOUND")
            return

        downloaded_ids: set[str] = set()
        for customer in customers:
            self.progress(f"BIN {bin_value}: customer {customer.name} ({customer.key})")
            dossier_html = self._post("/pkg_w_e_dossier.p_dossier_info", {
                "p_pcust_shortname": customer.key,
                "p_access": self.access,
            })
            dossiers = parse_dossiers(dossier_html)
            if dossier_number.strip():
                dossiers = [d for d in dossiers if dossier_matches(d, dossier_number)]
                if not dossiers:
                    self.progress(f"BIN {bin_value}: dossier № {dossier_number} not found")
                    self.log.add(bin=bin_value, customer=customer.name, dossier=dossier_number, result="DOSSIER NOT FOUND")
                    continue
                # BCC can return a parent "Общие документы" dossier with the same visible
                # number as the actual financial-leasing dossier. Prefer the specific dossier.
                specific = [d for d in dossiers if "общие документы" not in normalize_match_text(d.dossier_type)]
                if specific:
                    dossiers = specific
                self.progress(f"BIN {bin_value}: dossier filter matched {len(dossiers)} dossier(s)")
            if not dossiers:
                self.log.add(bin=bin_value, customer=customer.name, result="NO DOSSIER")
                continue

            for dossier in dossiers:
                self.progress(f"  dossier {dossier.dossier_id}: {dossier.label}")
                # Select dossier to mirror the site's normal state transition.
                self._post("/pkg_w_e_dossier.p_dossier_info", {
                    "p_pcust_shortname": customer.key,
                    "p_dossier": dossier.dossier_id,
                    "p_access": self.access,
                    "p_arh_status": "",
                })
                doc_info = self._post("/pkg_w_e_dossier.p_document_info", {
                    "p_dossier": dossier.dossier_id,
                    "p_pcust_shortname": customer.key,
                    "p_button": "change_dossier",
                    "p_access": self.access,
                    "p_arh_status": "",
                })
                architectures = parse_architectures(doc_info)
                if not architectures:
                    self.log.add(
                        bin=bin_value, customer=customer.name, dossier_id=dossier.dossier_id,
                        dossier=dossier.label, result="NO DOCUMENT CATEGORIES"
                    )
                    continue

                dossier_folder = bin_dir / safe_name(f"Dossier_{dossier.dossier_id}_{dossier.label}", 90)
                for architecture in architectures:
                    try:
                        detail = self._post("/PKG_W_E_DOSSIER.P_DOCUMENT_INFO_DETAIL", {
                            "p_dossier": dossier.dossier_id,
                            "p_section": architecture.section,
                            "p_access": self.access,
                            "p_architecture": architecture.architecture_id,
                            "p_arh_status": "",
                            "p_user_cnt": "0",
                            "p_centerproject_write": "0",
                        })
                        documents = parse_documents(detail, architecture)
                    except Exception as exc:
                        self.log.add(
                            bin=bin_value, customer=customer.name, dossier_id=dossier.dossier_id,
                            dossier=dossier.label, section=architecture.section,
                            architecture_id=architecture.architecture_id, category=architecture.label,
                            result="CATEGORY ERROR", error=str(exc)
                        )
                        continue

                    if not documents:
                        continue

                    target_documents = [d for d in documents if is_target_document(d)]
                    if not target_documents:
                        # Helpful diagnostic if BCC changes wording/encoding again.
                        titles = [clean_text(f"{d.category} | {d.title}") for d in documents[:3]]
                        if titles:
                            self.progress(f"    no target match in {architecture.label}; sample: {' ; '.join(titles)}")
                        continue

                    category_folder = dossier_folder / safe_name(
                        f"S{architecture.section}_{architecture.architecture_id}_{architecture.label}", 100
                    )
                    category_folder.mkdir(parents=True, exist_ok=True)
                    for document in target_documents:
                        if document.document_id in downloaded_ids:
                            continue
                        downloaded_ids.add(document.document_id)
                        self._download_document(
                            bin_value, customer, dossier, architecture, document, category_folder
                        )

        self.progress(
            f"BIN {bin_value}: completed; {len(downloaded_ids)} matching document(s) downloaded "
            f"[Заявление о присоединении (договор лизинга) / Договор купли-продажи]"
        )

    def _download_document(
        self, bin_value: str, customer: Customer, dossier: Dossier,
        architecture: Architecture, document: Document, folder: Path
    ):
        try:
            from urllib.parse import urlencode
            url = self.base + "/pkg_w_e_dossier.p_look_document?" + urlencode({
                "p_code_file": document.document_id,
                "p_name": "document_file",
            })
            result = self._browser_fetch(url, method="GET", binary=True)
            if not result.get("ok"):
                raise RuntimeError(f"HTTP {result.get('status')}")
            headers = {k.lower(): v for k, v in (result.get("headers") or {}).items()}
            content_type = headers.get("content-type", "")
            body = base64.b64decode(result.get("body_b64", ""))
            server_name = filename_from_content_disposition(headers.get("content-disposition", ""))
            suffix = Path(server_name).suffix if server_name else ""
            if not suffix:
                suffix = extension_for_content_type(content_type)
            base_name = safe_name(Path(server_name).stem if server_name else document.title, 135)
            file_name = f"{document.document_id}_{base_name}{suffix}"
            target = unique_path(folder / file_name)

            # Guard against login/error HTML being accidentally saved as a PDF.
            if "application/pdf" in content_type.lower() and not body.startswith(b"%PDF-"):
                raise RuntimeError("Server said PDF but response does not start with %PDF-; session may have expired")
            if "text/html" in content_type.lower() and body.lstrip().startswith((b"<", b"<!")):
                raise RuntimeError("Received HTML instead of a document; session may have expired or access is denied")

            target.write_bytes(body)
            self.progress(f"    saved {target.name}")
            self.log.add(
                bin=bin_value, customer=customer.name, dossier_id=dossier.dossier_id,
                dossier=dossier.label, section=architecture.section,
                architecture_id=architecture.architecture_id, category=document.category,
                document_id=document.document_id, document_title=document.title,
                document_number=document.number, document_date=document.date, status=document.status,
                result="DOWNLOADED", saved_path=str(target)
            )
        except Exception as exc:
            self.progress(f"    document {document.document_id}: ERROR: {exc}")
            self.log.add(
                bin=bin_value, customer=customer.name, dossier_id=dossier.dossier_id,
                dossier=dossier.label, section=architecture.section,
                architecture_id=architecture.architecture_id, category=document.category,
                document_id=document.document_id, document_title=document.title,
                document_number=document.number, document_date=document.date, status=document.status,
                result="DOWNLOAD ERROR", error=str(exc)
            )
